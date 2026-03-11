"""
╔══════════════════════════════════════════════════════════════════════════════╗
║   NEXUS ANALYZER PRO  8.0  ·  Pakistan Forensic CDR Intelligence           ║
║   Developed by: Faisal Malik                                                ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  ARCHITECTURE v8.0:                                                         ║
║  ┌─ NexusDashboard (QMainWindow) — Always visible main window              ║
║  │   Left : Create Case form + Cases list                                  ║
║  │   Right: Case overview cards + ROW BROWSER (all cases merged CDR)       ║
║  │   Footer: Developed by Faisal Malik                                     ║
║  │                                                                          ║
║  └─ CaseWindow (QMainWindow) — One window per case                         ║
║      Left sidebar: CDR Import + Crime Points (Cell ID / Date / Time)       ║
║      Right tabs:   Master | Common | A-Party | IMEI | Location |           ║
║                    Network | Direction | Timeline | Suspects |              ║
║                    Intel | DT-Analysis                                      ║
║      Bottom: Progress bar + Log + Footer                                   ║
║      Controls: Run Analysis / Export All / Export Dossier                  ║
║      Tower Browser + Date-Time Analysis → open as sub-windows              ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""
import sys, os, re, json, datetime, math, pickle, shutil
import pandas as pd
from PyQt5.QtWidgets import *
from PyQt5.QtCore    import (Qt, QDate, QTime, QThread, pyqtSignal,
                              QAbstractTableModel, QModelIndex, QTimer)
from PyQt5.QtGui     import QFont, QColor, QIcon, QBrush

# ─────────────────────────────────────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
BASE_PATH    = os.path.join(os.path.expanduser("~"), "Desktop", "NEXUS_Cases")
DISPLAY_COLS = ["A-Party","B-Party","Call Type","Duration","Date","Time","Day",
                "Cell ID","LAC","Location","Site","Network","Dir","IMEI","Hit"]
SHEET_KEYS   = ["1_Master","2_Common","3_AParty","4_IMEI","5_Location",
                "6_Network","7_Direction","8_Timeline","9_Suspects","10_Intelligence"]
SHEET_META = {
    "1_Master":("📋","Master","#1c2128","#58a6ff"),
    "2_Common":("🎯","Common","#0d2a1a","#3fb950"),
    "3_AParty":("📞","A-Party","#0d1e2e","#79c0ff"),
    "4_IMEI":("📱","IMEI","#2a1500","#e3b341"),
    "5_Location":("📡","Location","#1a0d2a","#d2a8ff"),
    "6_Network":("🌐","Network","#1c2128","#58a6ff"),
    "7_Direction":("↕","Direction","#0d1e2e","#79c0ff"),
    "8_Timeline":("📅","Timeline","#0d2a1a","#3fb950"),
    "9_Suspects":("🔴","Suspects","#3d0000","#f85149"),
    "10_Intelligence":("🧠","Intel","#1a0d2a","#d2a8ff"),
}
NET_COL = {"Jazz":"#e3b341","Zong":"#58a6ff","Telenor":"#3fb950",
           "Ufone":"#d2a8ff","SCO":"#f0883e"}

JAZZ_P={"0300","0301","0302","0303","0304","0305","0306","0307","0308","0309",
        "0320","0321","0322","0323","0324","0325","0326","0327","0328","0329"}
ZONG_P={"0310","0311","0312","0313","0314","0315","0316","0317","0318","0319"}
TEL_P ={"0340","0341","0342","0343","0344","0345","0346","0347","0348","0349"}
UFO_P ={"0330","0331","0332","0333","0334","0335","0336","0337","0338","0339"}
SCO_P ={"0855","0856","0857","0858","0859"}

# ─────────────────────────────────────────────────────────────────────────────
#  STYLESHEET
# ─────────────────────────────────────────────────────────────────────────────
SS = """
QWidget,QMainWindow,QDialog{background:#0d1117;color:#e6edf3;
  font-family:'Segoe UI',Consolas,Arial;font-size:12px;}
QGroupBox{border:1px solid #30363d;border-radius:6px;margin-top:8px;padding:8px;}
QGroupBox::title{color:#58a6ff;font-weight:bold;padding:0 6px;}
QLineEdit,QComboBox,QDateEdit,QTimeEdit,QSpinBox{background:#161b22;color:#e6edf3;
  border:1px solid #30363d;border-radius:4px;padding:3px 7px;min-height:26px;}
QLineEdit:focus,QDateEdit:focus,QComboBox:focus,QTimeEdit:focus{border-color:#58a6ff;}
QTextEdit{background:#161b22;color:#e6edf3;border:1px solid #30363d;
  border-radius:4px;padding:4px;}
QTableView,QTableWidget{background:#161b22;gridline-color:#21262d;color:#e6edf3;
  alternate-background-color:#0d1117;selection-background-color:#1f3a5a;}
QHeaderView::section{background:#1c2128;color:#58a6ff;font-weight:bold;
  padding:5px 8px;border:1px solid #21262d;}
QPushButton{background:#21262d;color:#e6edf3;border:1px solid #30363d;
  border-radius:5px;padding:4px 12px;}
QPushButton:hover{background:#30363d;border-color:#58a6ff;color:#58a6ff;}
QPushButton:pressed{background:#1f6feb;color:white;border-color:#1f6feb;}
QPushButton:disabled{color:#484f58;border-color:#21262d;background:#0d1117;}
QTabWidget::pane{border:1px solid #30363d;background:#0d1117;}
QTabBar::tab{background:#161b22;color:#8b949e;padding:6px 13px;
  border-radius:4px 4px 0 0;border:1px solid #30363d;border-bottom:none;
  margin-right:2px;}
QTabBar::tab:selected{background:#1c2128;color:#58a6ff;font-weight:bold;
  border-bottom:2px solid #58a6ff;}
QScrollBar:vertical{background:#161b22;width:6px;border-radius:3px;}
QScrollBar::handle:vertical{background:#30363d;border-radius:3px;min-height:20px;}
QScrollBar:horizontal{background:#161b22;height:6px;border-radius:3px;}
QScrollBar::handle:horizontal{background:#30363d;border-radius:3px;}
QProgressBar{border:none;background:#21262d;border-radius:2px;}
QProgressBar::chunk{background:#1f6feb;border-radius:2px;}
QListWidget{background:#161b22;border:1px solid #21262d;color:#e6edf3;}
QListWidget::item{padding:8px 10px;border-bottom:1px solid #21262d;}
QListWidget::item:selected{background:#1f3a5a;color:#58a6ff;}
QListWidget::item:hover{background:#21262d;}
QSplitter::handle{background:#21262d;}
QScrollArea{border:none;}
QLabel{color:#e6edf3;}
"""

_COL_W = {
    "#":40,"A-Party":115,"B-Party":115,"Call Type":82,"Duration":70,
    "Date":95,"Time":85,"Day":80,"Cell ID":90,"LAC":75,"Location":130,
    "Site":100,"Network":80,"Dir":88,"IMEI":140,"Hit":40,
    "At_Point":120,"Risk_Score":78,"Risk_Level":88,"Suspect_Flag":90,
    "MultiSIM":78,"Total_Calls":90,"Unique_B":80,"Cells_Used":85,
    "Unique_IMEI":90,"First_Date":95,"Last_Date":95,
    "INCOMING":82,"OUTGOING":82,"Case_Name":140,"Records":88,
    "Points":68,"Networks":110,"Last_Analyzed":130,"Status":90,
}

# ─────────────────────────────────────────────────────────────────────────────
#  NETWORK / CDR HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def detect_network(n):
    n=re.sub(r"\D","",str(n or ""))
    if not n: return "Unknown"
    if n.startswith("92") and len(n)==12: n="0"+n[2:]
    elif n.startswith("0092"): n="0"+n[4:]
    p=n[:4]
    if p in JAZZ_P: return "Jazz"
    if p in ZONG_P: return "Zong"
    if p in TEL_P:  return "Telenor"
    if p in UFO_P:  return "Ufone"
    if p in SCO_P:  return "SCO"
    if n[:3] in {"030","031","032"}: return "Jazz"
    if n[:3]=="034": return "Telenor"
    if n[:3]=="033": return "Ufone"
    return f"Unknown({n[:4]})"

def hex_to_dec(v):
    v=str(v or "").strip()
    if not v or v in ("nan","None","-",""): return v
    if re.match(r'^\d+$',v): return v
    if re.match(r'^[0-9A-Fa-f]{2,8}$',v):
        try: return str(int(v,16))
        except: pass
    return v

def _is_hex(v):
    v=str(v or "").strip()
    if not v or v.isdigit(): return False
    return bool(re.match(r'^[0-9A-Fa-f]{2,8}$',v))

def split_telenor(raw):
    s=str(raw or "").strip()
    if not s or s in ("nan","None","-"): return "",""
    for sep in ["-","_"," "]:
        if sep in s:
            p=s.split(sep,1); a,b=p[0].strip(),p[1].strip()
            if re.match(r'^\d+$',a) and re.match(r'^\d+$',b): return a,b
            return hex_to_dec(a),hex_to_dec(b)
    d=re.sub(r'\D','',s)
    if len(d)==8: return d[:4],d[4:]
    if len(d)==9: return d[:5],d[5:]
    if len(d)>=10:
        if len(d)==13: return d[5:9],d[9:]
        mid=len(d)//2; return d[:mid],d[mid:]
    return "",d

ALIASES={
    "A-Party":   ["a_party","a party","aparty","a-party","calling","calling number",
                  "caller","msisdn","number","originating","subscriber","cli","from",
                  "mobile number","a number","source number"],
    "B-Party":   ["b_party","b party","bparty","b-party","called","called number",
                  "dialled","dialed","destination","to","other party","b number",
                  "terminating number","dest number"],
    "Call Type": ["call_type","calltype","call type","type","service","event type",
                  "record type","category"],
    "Date":      ["date","call_date","calldate","call date","start_date","start date",
                  "event_date","dated","record date","transaction date"],
    "Time":      ["time","call_time","calltime","call time","start_time","start time",
                  "event_time","starttime"],
    "Cell ID":   ["cell id","cellid","cell_id","ci","cell","bts id","bts_id",
                  "btscode","cell identifier","combined cell","cell_no"],
    "LAC":       ["lac","location area code","location area","lac_id","area code"],
    "Location":  ["location","area","city","place","zone","region","address"],
    "Site":      ["site","site name","bts name","tower name","sitename","tower"],
    "Network":   ["network","operator","carrier","service provider","telecom"],
    "Dir":       ["dir","direction","call direction","mo_mt","type_dir"],
    "IMEI":      ["imei","device id","device imei","imei number","handset"],
    "Duration":  ["duration","dur","call duration","seconds","length","secs"],
}

def map_cols(df):
    cols={str(c).strip().lower():str(c).strip() for c in df.columns}
    rmap={}
    for tgt,aliases in ALIASES.items():
        for a in aliases:
            if a in cols: rmap[cols[a]]=tgt; break
    for raw_col,raw_raw in list(cols.items()):
        if "combined" in raw_col and "cell" in raw_col and "Cell ID" not in rmap.values():
            rmap[raw_raw]="Cell ID"
    return rmap

def normalize(raw,src=""):
    logs=[]; df=raw.copy(); df.columns=[str(c).strip() for c in df.columns]
    rmap=map_cols(df); df=df.rename(columns=rmap)
    for c in DISPLAY_COLS:
        if c not in df.columns: df[c]=""
    for col in ["A-Party","B-Party"]:
        s=df[col].astype(str).str.replace(r"\D","",regex=True).str.strip()
        df[col]=s.where(s!="nan","")
    unk=df["Network"].astype(str).str.strip().str.lower().isin(
        ["","nan","unknown","n/a","null","-","none","0"])
    if unk.any(): df.loc[unk,"Network"]=df.loc[unk,"A-Party"].apply(detect_network)
    sco_unk=df["Network"].astype(str).str.lower().isin(["unknown",""])
    if sco_unk.any():
        m=sco_unk&df["A-Party"].astype(str).str.startswith("085")
        if m.any(): df.loc[m,"Network"]="SCO"
    df["Network"]=df["Network"].astype(str).str.strip()
    net_col=df["Network"].str.strip().str.lower()
    cell_col=df["Cell ID"].astype(str); lac_col=df["LAC"].astype(str)
    tel=net_col=="telenor"
    if tel.any():
        sr=cell_col[tel].apply(split_telenor)
        lac_sp=sr.apply(lambda x:x[0]); cell_sp=sr.apply(lambda x:x[1])
        lac_empty=lac_col[tel].str.strip().isin(["","nan","None","-"])
        df.loc[tel,"Cell ID"]=cell_sp.values
        til=df.index[tel]; lev=lac_empty.values
        df.loc[til[lev],"LAC"]=lac_sp.values[lev]
        df.loc[tel,"Cell ID"]=df.loc[tel,"Cell ID"].astype(str).apply(hex_to_dec)
        df.loc[tel,"LAC"]=df.loc[tel,"LAC"].astype(str).apply(hex_to_dec)
    for net in ["jazz","ufone"]:
        m=net_col==net
        if m.any():
            df.loc[m,"LAC"]=lac_col[m].apply(hex_to_dec)
            df.loc[m,"Cell ID"]=cell_col[m].apply(lambda v:hex_to_dec(v) if _is_hex(v) else v)
    unk_net=~net_col.isin(["jazz","zong","telenor","ufone","sco"])
    if unk_net.any():
        df.loc[unk_net,"Cell ID"]=cell_col[unk_net].apply(lambda v:hex_to_dec(v) if _is_hex(v) else v)
        df.loc[unk_net,"LAC"]=lac_col[unk_net].apply(lambda v:hex_to_dec(v) if _is_hex(v) else v)
    for col in ["Cell ID","LAC"]:
        df[col]=(df[col].astype(str).str.replace(r'\.0$','',regex=True)
                        .str.strip().replace({"nan":"","None":"","none":"","-":""}))
    def vdate(s):
        out=pd.Series([""]*len(s),index=s.index); sv=s.astype(str).str.strip()
        for fmt in ["%Y-%m-%d","%d/%m/%Y","%m/%d/%Y","%d-%m-%Y","%Y/%m/%d","%d.%m.%Y","%Y%m%d"]:
            need=out==""
            if not need.any(): break
            p=pd.to_datetime(sv[need],format=fmt,errors="coerce"); hit=p.notna()
            if hit.any(): out[need]=p[hit].dt.strftime("%d/%m/%Y")
        remain=out==""
        if remain.any():
            p2=pd.to_datetime(sv[remain],dayfirst=True,errors="coerce"); h2=p2.notna()
            if h2.any(): out.loc[p2[h2].index]=p2[h2].dt.strftime("%d/%m/%Y")
        out[out.isin(["nan","None","NaT"])]=""; return out
    df["Date"]=vdate(df["Date"])
    def vtime(s):
        sv=s.astype(str).str.strip().str.replace(r'\.\d+$','',regex=True)
        ex=sv.str.extract(r'(\d{1,2}):(\d{2})(?::(\d{2}))?'); has=ex[0].notna(); out=sv.copy()
        if has.any():
            h=ex.loc[has,0].str.zfill(2); mn=ex.loc[has,1]; sc=ex.loc[has,2].fillna("00")
            out[has]=h+":"+mn+":"+sc
        out[out.isin(["nan","None","NaT",""])]=""; return out
    df["Time"]=vtime(df["Time"])
    pd_d=pd.to_datetime(df["Date"],format="%d/%m/%Y",errors="coerce")
    df["Day"]=pd_d.dt.strftime("%A").fillna("")
    df["Hit"]="1"; df["_src"]=src
    logs.append(f"  {src}: {len(df):,} rows — {dict(df['Network'].value_counts().head(4))}")
    return df,logs

# ─────────────────────────────────────────────────────────────────────────────
#  TABLE MODEL
# ─────────────────────────────────────────────────────────────────────────────
class PandasModel(QAbstractTableModel):
    def __init__(self):
        super().__init__(); self._df=pd.DataFrame()
    def load(self,df):
        self.beginResetModel(); self._df=df.reset_index(drop=True); self.endResetModel()
    def rowCount(self,p=QModelIndex()):    return len(self._df)
    def columnCount(self,p=QModelIndex()): return len(self._df.columns)
    def data(self,idx,role=Qt.DisplayRole):
        if not idx.isValid(): return None
        r,c=idx.row(),idx.column()
        if role==Qt.DisplayRole:
            v=self._df.iat[r,c]
            if v is None or v=="": return ""
            if type(v) is float:
                try:
                    if math.isnan(v): return ""
                except: pass
            return str(v)
        if role==Qt.TextAlignmentRole: return int(Qt.AlignCenter)
        if role==Qt.BackgroundRole and r%2==0: return QBrush(QColor("#0d1117"))
        return None
    def headerData(self,s,ori,role=Qt.DisplayRole):
        if role==Qt.DisplayRole:
            if ori==Qt.Horizontal:
                return str(self._df.columns[s]) if s<len(self._df.columns) else ""
            return str(s+1)
        return None

def make_tv():
    tv=QTableView()
    tv.setAlternatingRowColors(True)
    tv.setSelectionBehavior(QAbstractItemView.SelectRows)
    tv.setEditTriggers(QAbstractItemView.NoEditTriggers)
    tv.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
    tv.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
    tv.verticalHeader().setVisible(False)
    tv.setModel(PandasModel())
    tv.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
    tv.horizontalHeader().setStretchLastSection(True)
    tv.horizontalHeader().setMinimumSectionSize(40)
    tv.setWordWrap(False); return tv

def tv_load(tv,df):
    tv.model().load(df)
    hdr=tv.horizontalHeader()
    for i,col in enumerate(df.columns):
        hdr.resizeSection(i,_COL_W.get(str(col),110))

def pkl_save(df,path):
    try:
        with open(path,"wb") as f: pickle.dump(df,f,protocol=pickle.HIGHEST_PROTOCOL)
        return True
    except: return False

def pkl_load(path):
    try:
        with open(path,"rb") as f: df=pickle.load(f)
        return df if isinstance(df,pd.DataFrame) else None
    except: return None

# ─────────────────────────────────────────────────────────────────────────────
#  WORKERS
# ─────────────────────────────────────────────────────────────────────────────
class ImportWorker(QThread):
    progress=pyqtSignal(int,str); done=pyqtSignal(object,list,int); failed=pyqtSignal(str)
    def __init__(self,files,cache,raw_dir):
        super().__init__(); self.files=files; self.cache=cache; self.raw_dir=raw_dir
    def run(self):
        try:
            all_dfs,logs=[],[]; n=len(self.files); ok=0
            for i,fp in enumerate(self.files):
                fn=os.path.basename(fp); self.progress.emit(int(i/n*70),f"Reading: {fn}")
                try:
                    ext=os.path.splitext(fp)[1].lower(); raw=None
                    if ext==".csv":
                        for enc in ["utf-8","cp1252","latin-1","utf-16"]:
                            try: raw=pd.read_csv(fp,dtype=str,encoding=enc,on_bad_lines="skip"); break
                            except: pass
                    elif ext in [".xlsx",".xls"]: raw=pd.read_excel(fp,dtype=str)
                    elif ext==".txt":
                        for sep in ["\t",",","|",";"]:
                            try:
                                t=pd.read_csv(fp,sep=sep,dtype=str,on_bad_lines="skip")
                                if len(t.columns)>3: raw=t; break
                            except: pass
                    if raw is None: logs.append(f"Cannot read: {fn}"); continue
                    self.progress.emit(int((i+0.5)/n*70),f"Normalizing: {fn}")
                    dfn,lg=normalize(raw,fn); all_dfs.append(dfn); logs.extend(lg); ok+=1
                    dst=os.path.join(self.raw_dir,fn)
                    if not os.path.exists(dst):
                        try: shutil.copy2(fp,dst)
                        except: pass
                except Exception as e: logs.append(f"Error {fn}: {e}")
            if not all_dfs: self.failed.emit("Koi file load nahi ho saki."); return
            self.progress.emit(82,"Merging..."); combined=pd.concat(all_dfs,ignore_index=True)
            combined.insert(0,"#",range(1,len(combined)+1))
            self.progress.emit(90,"Saving cache..."); pkl_save(combined,self.cache)
            try:
                xl=os.path.join(os.path.dirname(self.cache),
                   os.path.basename(self.cache).replace(".pkl","_Master.xlsx"))
                self.progress.emit(95,"Saving Master.xlsx...")
                sc=[c for c in ["#"]+DISPLAY_COLS+["_src"] if c in combined.columns]
                combined[sc].to_excel(xl,index=False)
                logs.append(f"Master.xlsx saved")
            except Exception as e: logs.append(f"Master.xlsx warning: {e}")
            self.progress.emit(100,"Done!"); self.done.emit(combined,logs,ok)
        except Exception as e:
            import traceback; self.failed.emit(f"{e}\n{traceback.format_exc()[:400]}")

class CaseLoadWorker(QThread):
    progress=pyqtSignal(int,str); done=pyqtSignal(object); failed=pyqtSignal(str)
    def __init__(self,pkl_path,xl_path):
        super().__init__(); self.pkl_path=pkl_path; self.xl_path=xl_path
    def run(self):
        try:
            if os.path.exists(self.pkl_path):
                self.progress.emit(40,"Loading cache...")
                df=pkl_load(self.pkl_path)
                if df is not None: self.progress.emit(100,"Loaded!"); self.done.emit(df); return
            if os.path.exists(self.xl_path):
                self.progress.emit(40,"Reading Excel...")
                df=pd.read_excel(self.xl_path,dtype=str)
                self.progress.emit(85,"Caching..."); pkl_save(df,self.pkl_path)
                self.progress.emit(100,"Done!"); self.done.emit(df); return
            self.done.emit(pd.DataFrame())
        except Exception as e: self.failed.emit(str(e))

class SaveWorker(QThread):
    progress=pyqtSignal(int,str); done=pyqtSignal(str); failed=pyqtSignal(str)
    def __init__(self,tasks):
        super().__init__(); self.tasks=tasks
    def run(self):
        try:
            n=max(len(self.tasks),1)
            for i,(df,fp) in enumerate(self.tasks):
                self.progress.emit(int(i/n*90)+5,f"Saving {os.path.basename(fp)}...")
                df.to_excel(fp,index=False)
            self.progress.emit(100,"Done!"); self.done.emit(f"{len(self.tasks)} saved")
        except Exception as e:
            import traceback; self.failed.emit(f"{e}\n{traceback.format_exc()[:300]}")

class MultiSheetExportWorker(QThread):
    progress=pyqtSignal(int,str); done=pyqtSignal(str); failed=pyqtSignal(str)
    def __init__(self,dfs_map,filepath):
        super().__init__(); self.dfs_map=dfs_map; self.filepath=filepath
    def run(self):
        try:
            from openpyxl.styles import PatternFill,Font,Alignment,Border,Side
            from openpyxl.utils import get_column_letter
            w=pd.ExcelWriter(self.filepath,engine="openpyxl")
            keys=list(self.dfs_map.keys()); n=max(len(keys),1)
            for i,(sname,df) in enumerate(self.dfs_map.items()):
                self.progress.emit(int(5+i/n*65),f"Writing {sname}...")
                df.to_excel(w,sheet_name=sname[:31],index=False)
            self.progress.emit(75,"Styling...")
            hf=PatternFill("solid",fgColor="0D1117")
            hft=Font(bold=True,color="58A6FF",size=10)
            af=PatternFill("solid",fgColor="161B22")
            bs=Side(style="thin",color="21262D"); bdr=Border(left=bs,right=bs,top=bs,bottom=bs)
            for ws in w.book.worksheets:
                for cell in ws[1]:
                    cell.fill=hf; cell.font=hft
                    cell.alignment=Alignment(horizontal="center",wrap_text=True)
                    cell.border=bdr
                for ri,row in enumerate(ws.iter_rows(min_row=2),2):
                    fill=af if ri%2==0 else PatternFill()
                    for cell in row:
                        cell.fill=fill; cell.border=bdr
                        cell.alignment=Alignment(horizontal="left",vertical="center")
                ws.freeze_panes="A2"
                for ci,col in enumerate(ws.columns,1):
                    ml=max((len(str(c.value or "")) for c in list(col)[:200]),default=8)
                    ws.column_dimensions[get_column_letter(ci)].width=min(ml+4,42)
            self.progress.emit(95,"Saving..."); w.close()
            self.progress.emit(100,"Done!"); self.done.emit(f"Saved: {os.path.basename(self.filepath)}")
        except Exception as e:
            import traceback; self.failed.emit(f"{e}\n{traceback.format_exc()[:300]}")

class DossierWorker(QThread):
    progress=pyqtSignal(int,str); done=pyqtSignal(str); failed=pyqtSignal(str)
    def __init__(self,df,meta,common,multi_im,filepath):
        super().__init__(); self.df=df; self.meta=meta; self.common=common
        self.multi_im=multi_im; self.filepath=filepath
    def run(self):
        try:
            self.progress.emit(10,"Writing Dossier...")
            _write_dossier(self.df,self.meta,self.common,self.multi_im,self.filepath)
            self.progress.emit(100,"Done!"); self.done.emit(f"Dossier: {os.path.basename(self.filepath)}")
        except Exception as e:
            import traceback; self.failed.emit(f"{e}\n{traceback.format_exc()[:300]}")

class AnalysisWorker(QThread):
    progress=pyqtSignal(int,str); sheet=pyqtSignal(str,object)
    done=pyqtSignal(int,object,object)   # suspect_count, common set, multi_im set
    failed=pyqtSignal(str)
    def __init__(self,df,meta):
        super().__init__(); self.df=df; self.meta=meta
    def _emit(self,key,df,pct):
        self.sheet.emit(key,df); self.progress.emit(pct,"")
    def run(self):
        try:
            df=self.df; meta=self.meta
            for c in DISPLAY_COLS:
                if c not in df.columns: df[c]=""
            disp=[c for c in DISPLAY_COLS if c in df.columns]
            ddt=pd.to_datetime(df["Date"],format="%d/%m/%Y",errors="coerce")
            # ── Point sets ──────────────────────────────────────────────────
            pts=meta.get("points",[]); pt_labels={}; pt_set={}
            for pt in pts:
                cells=set(str(c).strip() for c in pt.get("cells",[]))
                for c in cells: pt_labels[c]=f"{pt.get('letter','?')}: {pt.get('name','?')}"
                mask=df["Cell ID"].astype(str).str.strip().isin(cells)
                try:
                    df_from=pd.to_datetime(pt["dt_from"]); df_to=pd.to_datetime(pt["dt_to"])
                    ddt_full=pd.to_datetime(df["Date"]+" "+df["Time"].fillna("00:00:00"),
                                           format="%d/%m/%Y %H:%M:%S",errors="coerce")
                    mask&=(ddt_full>=df_from)&(ddt_full<=df_to)
                except: pass
                pt_set[pt["letter"]]=set(df.loc[mask,"A-Party"].dropna().astype(str).unique())
            common=(set.intersection(*pt_set.values()) if len(pt_set)>1
                    else list(pt_set.values())[0] if pt_set else set())
            # ── Sheet 1 Master ───────────────────────────────────────────────
            self.progress.emit(5,"Master CDR...")
            self._emit("1_Master",df[disp],10)
            # ── Sheet 2 Common ───────────────────────────────────────────────
            self.progress.emit(12,"Common Targets...")
            cm=df[df["A-Party"].isin(common)][disp].copy()
            if not cm.empty and "Cell ID" in cm.columns:
                cm.insert(0,"At_Point",cm["Cell ID"].astype(str).str.strip().map(pt_labels).fillna("—"))
            self._emit("2_Common",cm,20)
            # ── Sheet 3 A-Party ──────────────────────────────────────────────
            self.progress.emit(22,"A-Party Summary...")
            g3=df.groupby("A-Party",dropna=False,sort=False)
            ap=pd.concat([g3["B-Party"].count().rename("Total_Calls"),
                          g3["B-Party"].nunique().rename("Unique_B"),
                          g3["Cell ID"].nunique().rename("Cells_Used"),
                          g3["IMEI"].nunique().rename("Unique_IMEI"),
                          g3["Date"].min().rename("First_Date"),
                          g3["Date"].max().rename("Last_Date")],axis=1).reset_index()
            ap["Network"]=df.groupby("A-Party",sort=False)["Network"].first().reindex(ap["A-Party"].values).values
            dir_norm=df["Dir"].astype(str).str.strip().str.upper().replace(
                {"INCOMING":"INCOMING","INBOUND":"INCOMING","MTC":"INCOMING","MT":"INCOMING","IN":"INCOMING",
                 "OUTGOING":"OUTGOING","OUTBOUND":"OUTGOING","MOC":"OUTGOING","MO":"OUTGOING","OUT":"OUTGOING"})
            _tmp=df.assign(_Dir=dir_norm)
            dg=_tmp.groupby(["A-Party","_Dir"],sort=False).size().unstack(fill_value=0)
            for d in ["INCOMING","OUTGOING"]:
                ap[d]=ap["A-Party"].map(dg[d] if d in dg.columns else pd.Series(dtype=int)).fillna(0).astype(int)
            mx_b=ap["Unique_B"].max() or 1; mx_c=ap["Cells_Used"].max() or 1
            ap["Risk_Score"]=((ap["Unique_IMEI"]>1).astype(int)*40+
                              (ap["Unique_B"]/mx_b*30).round(0).astype(int)+
                              (ap["Cells_Used"]/mx_c*30).round(0).astype(int)).clip(0,100)
            ap["Risk_Level"]=pd.cut(ap["Risk_Score"],bins=[-1,25,50,75,101],
                labels=["Low","Medium","High","Critical"])
            ap["Suspect_Flag"]=ap["A-Party"].isin(common).map({True:"YES",False:"—"})
            ap["MultiSIM"]=(ap["Unique_IMEI"]>1).map({True:"YES",False:"—"})
            ap=ap.sort_values("Risk_Score",ascending=False).reset_index(drop=True)
            self._emit("3_AParty",ap,32)
            # ── Sheet 4 IMEI ─────────────────────────────────────────────────
            self.progress.emit(34,"IMEI Analysis...")
            g4=df.groupby("IMEI",dropna=False,sort=False)
            im=pd.concat([g4["A-Party"].nunique().rename("SIMs_Used"),
                          g4["B-Party"].count().rename("Total_Calls"),
                          g4["Cell ID"].nunique().rename("Towers_Used"),
                          g4["Date"].min().rename("First_Seen"),
                          g4["Date"].max().rename("Last_Seen")],axis=1).reset_index()
            top_num=(df.groupby(["IMEI","A-Party"],sort=False).size().reset_index(name="_n")
                     .sort_values("_n",ascending=False).drop_duplicates("IMEI")
                     .set_index("IMEI")["A-Party"])
            im["Primary_Number"]=im["IMEI"].map(top_num).fillna("")
            im["Suspect_IMEI"]=im["Primary_Number"].isin(common).map({True:"YES",False:"—"})
            im["Risk_Level"]=im["SIMs_Used"].apply(
                lambda s:"Critical" if s>3 else "High" if s>1 else "Low")
            multi_im=set(im[im["SIMs_Used"]>1]["IMEI"].dropna().astype(str))
            im=im.sort_values(["SIMs_Used","Total_Calls"],ascending=False).reset_index(drop=True)
            self._emit("4_IMEI",im,42)
            # ── Sheet 5 Location ─────────────────────────────────────────────
            self.progress.emit(44,"Location HeatMap...")
            g5c=[c for c in ["Cell ID","LAC","Location","Site","Network"] if c in df.columns]
            g5=df.groupby(g5c,dropna=False,sort=False)
            lc=pd.concat([g5.size().rename("Total_Hits"),
                          g5["A-Party"].nunique().rename("Unique_Numbers"),
                          g5["Date"].nunique().rename("Active_Days")],axis=1).reset_index()
            sc2=df[df["A-Party"].isin(common)].groupby("Cell ID",sort=False)["A-Party"].nunique()
            lc["Suspect_Count"]=lc["Cell ID"].map(sc2).fillna(0).astype(int)
            lc["Point_Label"]=lc["Cell ID"].astype(str).map(pt_labels).fillna("—")
            lc["Risk_Level"]=lc.apply(lambda r:"Critical" if r["Suspect_Count"]>=3 else
                "High" if r["Suspect_Count"]>=1 else "Medium" if r["Unique_Numbers"]>50 else "Low",axis=1)
            lc=lc.sort_values(["Suspect_Count","Total_Hits"],ascending=False).reset_index(drop=True)
            self._emit("5_Location",lc,52)
            # ── Sheet 6 Network ──────────────────────────────────────────────
            self.progress.emit(54,"Network Breakdown...")
            g6=df.groupby("Network",dropna=False,sort=False)
            nt=pd.concat([g6.size().rename("Total_Records"),
                          g6["A-Party"].nunique().rename("Unique_A"),
                          g6["B-Party"].nunique().rename("Unique_B"),
                          g6["Cell ID"].nunique().rename("Towers"),
                          g6["Date"].nunique().rename("Active_Days")],axis=1).reset_index()
            tr=nt["Total_Records"].sum() or 1
            nt["Share_%"]=(nt["Total_Records"]/tr*100).round(1)
            nt=nt.sort_values("Total_Records",ascending=False).reset_index(drop=True)
            self._emit("6_Network",nt,61)
            # ── Sheet 7 Direction ────────────────────────────────────────────
            self.progress.emit(63,"Call Direction...")
            _ds=df["Dir"].astype(str).str.strip().str.upper().replace(
                {"INCOMING":"INCOMING","INBOUND":"INCOMING","MTC":"INCOMING","MT":"INCOMING","IN":"INCOMING",
                 "OUTGOING":"OUTGOING","OUTBOUND":"OUTGOING","MOC":"OUTGOING","MO":"OUTGOING","OUT":"OUTGOING"})
            try:
                dr=(df.assign(_Dir=_ds).groupby(["A-Party","_Dir"],dropna=False,sort=False)
                   .size().unstack(fill_value=0).reset_index()); dr.columns.name=None
                dc=[c for c in dr.columns if c!="A-Party"]; dr["Total_Calls"]=dr[dc].sum(axis=1)
                if "INCOMING" in dr.columns and "OUTGOING" in dr.columns:
                    dr["IN_%"]=(dr["INCOMING"]/dr["Total_Calls"].replace(0,1)*100).round(1)
                    dr["Pattern"]=dr.apply(lambda r:
                        "Receive-Only" if r.get("OUTGOING",0)==0 else
                        "Call-Only" if r.get("INCOMING",0)==0 else "Bidirectional",axis=1)
                dr["Suspect_Flag"]=dr["A-Party"].isin(common).map({True:"YES",False:"—"})
                dr=dr.sort_values("Total_Calls",ascending=False).reset_index(drop=True)
            except: dr=pd.DataFrame(columns=["A-Party"])
            self._emit("7_Direction",dr,70)
            # ── Sheet 8 Timeline ─────────────────────────────────────────────
            self.progress.emit(72,"Daily Timeline...")
            tg=[c for c in ["Date","Day"] if c in df.columns]
            g8=df.groupby(tg,dropna=False,sort=False)
            tl=pd.concat([g8.size().rename("Total_Records"),
                          g8["A-Party"].nunique().rename("Active_Numbers"),
                          g8["B-Party"].nunique().rename("Unique_Contacts"),
                          g8["Cell ID"].nunique().rename("Towers_Used")],axis=1).reset_index()
            rm=tl["Total_Records"].mean(); rs=tl["Total_Records"].std() or 1
            tl["Spike_Flag"]=tl["Total_Records"].apply(lambda x:"SPIKE" if x>rm+2*rs else "—")
            tl["_s"]=pd.to_datetime(tl["Date"],format="%d/%m/%Y",errors="coerce")
            tl=tl.sort_values("_s").drop(columns=["_s"]).reset_index(drop=True)
            self._emit("8_Timeline",tl,79)
            # ── Sheet 9 Suspects ─────────────────────────────────────────────
            self.progress.emit(81,"Suspects...")
            smask=df["A-Party"].isin(common)|df["IMEI"].astype(str).isin(multi_im)
            sd=df.loc[smask,disp].copy()
            is_c=df.loc[smask,"A-Party"].isin(common)
            is_m=df.loc[smask,"IMEI"].astype(str).isin(multi_im)
            reasons=(is_c.map({True:"Common-Target",False:""})+
                     is_m.map({True:", Multi-SIM",False:""})).str.strip(", ")
            reasons=reasons.where(reasons!="","Suspect")
            sd.insert(0,"Reason",reasons.values)
            if not sd.empty:
                sg=sd.groupby("A-Party",sort=False)
                def _pts_for(num):
                    found=[l for l,nums in pt_set.items() if num in nums]
                    return ",".join(sorted(found)) if found else "—"
                pts_map={n:_pts_for(n) for n in sd["A-Party"].dropna().unique()}
                sd.insert(1,"Points_At",sd["A-Party"].map(pts_map).values)
                sd.insert(2,"IMEI_Count",sd["A-Party"].map(sg["IMEI"].nunique()).values)
                sd.insert(3,"Total_Calls",sd["A-Party"].map(sg.size()).values)
                sd.insert(4,"Unique_Contacts",sd["A-Party"].map(sg["B-Party"].nunique()).values)
                sd.insert(5,"Towers_Used",sd["A-Party"].map(sg["Cell ID"].nunique()).values)
            self._emit("9_Suspects",sd,89)
            # ── Sheet 10 Intelligence ────────────────────────────────────────
            self.progress.emit(91,"Intelligence Report...")
            intel=_build_intel(df,meta,common,multi_im,pt_labels)
            self._emit("10_Intelligence",intel,97)
            self.progress.emit(100,"Analysis complete!")
            self.done.emit(len(common),common,multi_im)
        except Exception as e:
            import traceback; self.failed.emit(f"{e}\n{traceback.format_exc()[:600]}")

class DTAnalysisWorker(QThread):
    progress=pyqtSignal(int,str); done=pyqtSignal(object); failed=pyqtSignal(str)
    def __init__(self,df,dt_from,dt_to,tm_from="00:00:00",tm_to="23:59:59",filters=None):
        super().__init__(); self.df=df; self.dt_from=dt_from; self.dt_to=dt_to
        self.tm_from=tm_from; self.tm_to=tm_to; self.filters=filters or {}
    def run(self):
        try:
            df=self.df; self.progress.emit(10,"Filtering by date/time...")
            ddt=pd.to_datetime(df["Date"],format="%d/%m/%Y",errors="coerce")
            mask=(ddt>=pd.to_datetime(self.dt_from))&(ddt<=pd.to_datetime(self.dt_to))
            if self.tm_from!="00:00:00" or self.tm_to!="23:59:59":
                try:
                    tc=df["Time"].astype(str).str[:8].str.strip()
                    mask&=(tc>=self.tm_from)&(tc<=self.tm_to)
                except: pass
            for col,val in self.filters.items():
                if val and col in df.columns:
                    mask&=df[col].astype(str).str.lower().str.contains(
                        str(val).lower(),regex=False,na=False)
            filtered=df.loc[mask].copy()
            if filtered.empty:
                self.done.emit(pd.DataFrame([{"Info":"No records in range."}])); return
            self.progress.emit(40,"Stats per number...")
            g=filtered.groupby("A-Party",dropna=False,sort=False)
            result=pd.concat([g.size().rename("Total_Calls"),
                              g["Date"].nunique().rename("Active_Days"),
                              g["B-Party"].nunique().rename("Unique_Contacts"),
                              g["IMEI"].nunique().rename("IMEI_Count"),
                              g["Cell ID"].nunique().rename("Towers_Used"),
                              g["Date"].min().rename("First_Date"),
                              g["Date"].max().rename("Last_Date"),
                              g["Network"].first().rename("Network")],axis=1).reset_index()
            result=result.sort_values(["Active_Days","Total_Calls"],ascending=False).reset_index(drop=True)
            self.progress.emit(65,"IMEI cross-match...")
            imei_sims=filtered.groupby("IMEI",dropna=False)["A-Party"].nunique()
            multi_imei=set(imei_sims[imei_sims>1].index.astype(str))
            result["Multi_IMEI"]=result["IMEI_Count"].apply(lambda x:"YES" if x>1 else "—")
            result["Shared_Device"]=result["A-Party"].apply(lambda n:"YES" if any(
                str(im) in multi_imei for im in filtered[filtered["A-Party"]==n]["IMEI"].dropna().unique()) else "—")
            self.progress.emit(82,"ABC/CBA mutual calls...")
            calls_ab=set(zip(filtered["A-Party"].astype(str),filtered["B-Party"].astype(str)))
            mutual=set()
            for a,b in calls_ab:
                if (b,a) in calls_ab: mutual.add(a); mutual.add(b)
            result["ABC_CBA"]=result["A-Party"].astype(str).isin(mutual).map({True:"MUTUAL",False:"—"})
            result["Multi_Date"]=result["Active_Days"].apply(lambda x:"YES" if x>=2 else "—")
            top_c={}
            for num,grp in filtered.groupby("A-Party",sort=False):
                bc=grp["B-Party"].value_counts()
                if len(bc): top_c[num]=f"{bc.index[0]}({bc.iloc[0]}x)"
            result["Top_Contact"]=result["A-Party"].map(top_c).fillna("—")
            main_cell=filtered.groupby("A-Party",sort=False)["Cell ID"].agg(
                lambda x:x.mode().iloc[0] if len(x) else "")
            result["Main_Cell"]=result["A-Party"].map(main_cell).fillna("—")
            result.insert(0,"Range_Records",len(filtered))
            self.progress.emit(100,"Done!"); self.done.emit(result)
        except Exception as e:
            import traceback; self.failed.emit(f"{e}\n{traceback.format_exc()[:400]}")

def _build_intel(df,meta,common,multi_im,pt_labels):
    grp=df.groupby(["A-Party","B-Party"],dropna=False,sort=False)
    intel=pd.concat([grp.size().rename("Calls"),grp["Date"].min().rename("First"),
                     grp["Date"].max().rename("Last")],axis=1).reset_index()
    intel=intel.sort_values("Calls",ascending=False)
    net_a=df.groupby("A-Party",sort=False)["Network"].first()
    net_b=df.groupby("B-Party",sort=False)["Network"].first()
    tower_m=df.groupby("A-Party",sort=False)["Cell ID"].agg(
        lambda x:x.mode().iloc[0] if len(x)>0 else "")
    intel["A_Net"]=intel["A-Party"].map(net_a).fillna("Unknown")
    intel["B_Net"]=intel["B-Party"].map(net_b).fillna("Unknown")
    intel["A_Suspect"]=intel["A-Party"].isin(common).map({True:"YES",False:"—"})
    intel["B_Suspect"]=intel["B-Party"].isin(common).map({True:"YES",False:"—"})
    intel["A_MultiSIM"]=intel["A-Party"].apply(lambda x:"YES" if str(x) in multi_im else "—")
    intel["Main_Tower"]=intel["A-Party"].map(tower_m).fillna("")
    intel["Tower_Point"]=intel["Main_Tower"].map(pt_labels).fillna("—")
    def conn(row):
        a=row["A_Suspect"]=="YES"; b=row["B_Suspect"]=="YES"
        if a and b: return "Suspect <-> Suspect"
        if a: return "Suspect -> Contact"
        if b: return "Contact -> Suspect"
        return "Background"
    intel["Link_Type"]=intel.apply(conn,axis=1)
    cols=["Link_Type","A-Party","A_Net","A_Suspect","A_MultiSIM","B-Party","B_Net",
          "B_Suspect","Calls","First","Last","Main_Tower","Tower_Point"]
    return intel[[c for c in cols if c in intel.columns]]

def _write_dossier(df,meta,common,multi_im,filepath):
    try: from openpyxl.styles import PatternFill,Font,Alignment,Border,Side
    except: return
    suspects=sorted(common); pts=meta.get("points",[])
    with pd.ExcelWriter(filepath,engine="openpyxl") as w:
        # Sheet 1 Summary
        rows=[["NEXUS ANALYZER PRO 8.0 — INTELLIGENCE DOSSIER",""],
              ["Developed by: Faisal Malik",""],
              ["Case Name",meta.get("case_name","")],
              ["Total Records",len(df)],["Suspects Found",len(common)],
              ["Date Range",f"{df['Date'].min()} to {df['Date'].max()}"],
              ["Points",len(pts)],
              ["Exported",datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
              ["",""],["SUSPECTS",""]]
        for n in suspects: rows.append([n,""])
        pd.DataFrame(rows).to_excel(w,sheet_name="1_Case_Summary",index=False,header=False)
        # Sheet 2 Profiles
        profiles=[]
        for num in suspects[:300]:
            sub=df[df["A-Party"]==num]
            if sub.empty: continue
            profiles.append({"Number":num,"Network":sub["Network"].mode().iloc[0] if len(sub) else "",
                "Total_Calls":len(sub),"Unique_Contacts":sub["B-Party"].nunique(),
                "IMEI_Count":sub["IMEI"].nunique(),"Towers":sub["Cell ID"].nunique(),
                "First":sub["Date"].min(),"Last":sub["Date"].max(),
                "IN":(sub["Dir"].astype(str).str.upper().isin(["INCOMING","MTC","MT","IN"])).sum(),
                "OUT":(sub["Dir"].astype(str).str.upper().isin(["OUTGOING","MOC","MO","OUT"])).sum()})
        pd.DataFrame(profiles).to_excel(w,sheet_name="2_Suspect_Profiles",index=False)
        # Sheet 3 Matrix
        top_c=(df[df["A-Party"].isin(common)]["B-Party"].value_counts().head(30).index.tolist())
        mrows=[]
        for num in suspects[:50]:
            sub=df[df["A-Party"]==num]; row={"Suspect":num}
            for c in top_c: row[c]=len(sub[sub["B-Party"]==c])
            mrows.append(row)
        if mrows: pd.DataFrame(mrows).to_excel(w,sheet_name="3_Connection_Matrix",index=False)
        # Sheet 4 Relationships
        pairs=[]
        for num in suspects[:100]:
            sub=df[df["A-Party"]==num]
            for b,grp in sub.groupby("B-Party"):
                if not b or str(b)=="nan": continue
                calls=len(grp); days=grp["Date"].nunique()
                score=calls*2+days*5
                level="CRITICAL" if score>200 else "HIGH" if score>80 else "MEDIUM" if score>30 else "LOW"
                pairs.append({"A-Party":num,"B-Party":b,"Calls":calls,"Days":days,
                    "Score":score,"Level":level,"B_Suspect":"YES" if b in common else "—"})
        if pairs: pd.DataFrame(pairs).sort_values("Score",ascending=False).to_excel(
            w,sheet_name="4_Relationship_Strength",index=False)
        # Sheet 5 Tower CoPresence
        cop=[]
        for pt in pts:
            cells=set(str(c).strip() for c in pt.get("cells",[]))
            pt_df=df[df["Cell ID"].astype(str).str.strip().isin(cells)].head(3000)
            for _,row in pt_df.iterrows():
                cop.append({"Point":f"{pt.get('letter','?')}: {pt.get('name','?')}",
                    "Cell ID":row.get("Cell ID",""),"Date":row.get("Date",""),
                    "A-Party":row.get("A-Party",""),"Network":row.get("Network",""),
                    "Suspect":"YES" if row.get("A-Party","") in common else "—"})
        if cop: pd.DataFrame(cop).to_excel(w,sheet_name="5_Tower_CoPresence",index=False)
        # Sheet 6 Hourly Heatmap
        try:
            df2=df.copy()
            df2["_hr"]=pd.to_numeric(df2["Time"].astype(str).str[:2].str.replace(r"\D","",regex=True),errors="coerce")
            heat=df2[df2["A-Party"].isin(common)].groupby(["_hr","A-Party"],sort=False).size().unstack(fill_value=0)
            heat.index=[f"{int(h):02d}:00" if str(h).replace(".0","").isdigit() else str(h) for h in heat.index]
            heat.reset_index().rename(columns={"_hr":"Hour"}).to_excel(w,sheet_name="6_Hourly_Heatmap",index=False)
        except: pass
        # Sheet 7 Point Presence
        for pt in pts:
            cells=set(str(c).strip() for c in pt.get("cells",[]))
            sn=f"7_Pt{pt.get('letter','?')}_{pt.get('name','?')}"[:31]
            pt_df=df[df["Cell ID"].astype(str).str.strip().isin(cells)].copy().head(5000)
            if not pt_df.empty:
                pt_df=pt_df.copy()
                pt_df.insert(0,"Suspect",pt_df["A-Party"].isin(common).map({True:"YES",False:"—"}))
                dc=[c for c in ["Suspect"]+DISPLAY_COLS if c in pt_df.columns]
                pt_df[dc].to_excel(w,sheet_name=sn,index=False)

# ─────────────────────────────────────────────────────────────────────────────
#  SHARED WIDGET HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _b(txt,col="#21262d",fn=None,h=30,bold=False,w=None,parent=None):
    b=QPushButton(txt,parent); b.setFixedHeight(h)
    if w: b.setFixedWidth(w)
    s=f"background:{col};color:white;border-radius:5px;padding:3px 10px;border:none;"
    if bold: s+="font-weight:bold;"
    b.setStyleSheet(s)
    if fn: b.clicked.connect(fn)
    return b

def _l(parent_or_txt,txt_or_col="#e6edf3",col_or_sz="#e6edf3",sz_or_bold=12,bold_or_none=False,_extra=None):
    # Supports both _l(parent,txt,col,sz,bold) and _l(txt,col,sz,bold,parent)
    from PyQt5.QtWidgets import QWidget
    if isinstance(parent_or_txt, QWidget):
        parent,txt,col,sz,bold = parent_or_txt,txt_or_col,col_or_sz,sz_or_bold,bold_or_none
    else:
        txt,col,sz,bold,parent = parent_or_txt,txt_or_col,col_or_sz,sz_or_bold,bold_or_none
    lb=QLabel(str(txt),parent); s=f"color:{col};font-size:{sz}px;"
    if bold: s+="font-weight:bold;"
    lb.setStyleSheet(s); return lb

def _sep(parent=None):
    f=QFrame(parent); f.setFrameShape(QFrame.HLine)
    f.setStyleSheet("background:#21262d;border:none;max-height:1px;"); return f

def _chip_widget(parent,label,val,col):
    f=QFrame(parent); f.setStyleSheet("QFrame{background:transparent;border:none;}")
    fl=QHBoxLayout(f); fl.setContentsMargins(0,0,0,0); fl.setSpacing(3)
    v=QLabel(val); v.setStyleSheet(f"color:{col};font-weight:bold;font-size:12px;")
    lb=QLabel(label); lb.setStyleSheet("color:#8b949e;font-size:9px;")
    fl.addWidget(v); fl.addWidget(lb); f._v=v; return f

def _footer_widget(parent=None):
    ft=QWidget(parent); ft.setFixedHeight(22)
    ft.setStyleSheet("background:#161b22;border-top:1px solid #21262d;")
    ftl=QHBoxLayout(ft); ftl.setContentsMargins(12,0,12,0)
    ftl.addWidget(_l(ft,"NEXUS ANALYZER PRO 8.0","#58a6ff",9,True))
    ftl.addStretch()
    ftl.addWidget(_l(ft,"Developed by Faisal Malik","#3fb950",9,True))
    ftl.addStretch()
    ftl.addWidget(_l(ft,"Pakistan Forensic CDR Intelligence","#8b949e",9))
    return ft

# ─────────────────────────────────────────────────────────────────────────────
#  POINT CARD WIDGET
# ─────────────────────────────────────────────────────────────────────────────
class PointCard(QFrame):
    sig_remove=pyqtSignal(object)
    sig_browse=pyqtSignal(object)
    def __init__(self,letter,data=None,locked=False):
        super().__init__(); self.letter=letter; self.setObjectName("PC")
        g=QGridLayout(self); g.setContentsMargins(8,6,8,6); g.setSpacing(4)
        self.lbl=QLabel(f"Point {letter}")
        self.lbl.setStyleSheet("color:#f0883e;font-weight:bold;font-size:12px;")
        g.addWidget(self.lbl,0,0,1,3)
        bx=QPushButton("✕"); bx.setFixedSize(20,20)
        bx.setStyleSheet("background:#6e1a1a;color:#fca5a5;border-radius:3px;border:none;font-size:9px;")
        bx.clicked.connect(lambda:self.sig_remove.emit(self)); g.addWidget(bx,0,3,Qt.AlignRight)
        g.addWidget(QLabel("Name:"),1,0)
        self.e_name=QLineEdit(data.get("name","") if data else "")
        self.e_name.setPlaceholderText("e.g. Crime Scene A"); g.addWidget(self.e_name,1,1,1,3)
        g.addWidget(QLabel("Cells:"),2,0)
        self.e_cells=QTextEdit(); self.e_cells.setFixedHeight(46)
        self.e_cells.setPlaceholderText("Cell IDs (one per line)\nUse Browse to pick from Master")
        if data and data.get("cells"): self.e_cells.setPlainText("\n".join(data["cells"]))
        g.addWidget(self.e_cells,2,1,1,2)
        brw=QPushButton("Browse"); brw.setFixedSize(52,46)
        brw.setStyleSheet("background:#0d2a3a;color:#58a6ff;border:1px solid #1f6feb;"
                          "border-radius:5px;font-size:10px;font-weight:bold;")
        brw.clicked.connect(lambda:self.sig_browse.emit(self)); g.addWidget(brw,2,3)
        g.addWidget(QLabel("Date:"),3,0)
        dr=QHBoxLayout(); dr.setSpacing(3)
        self.d_from=QDateEdit(); self.d_from.setCalendarPopup(True); self.d_from.setDisplayFormat("dd/MM/yyyy")
        self.d_to=QDateEdit(); self.d_to.setCalendarPopup(True); self.d_to.setDisplayFormat("dd/MM/yyyy")
        dr.addWidget(QLabel("From:")); dr.addWidget(self.d_from)
        dr.addWidget(QLabel("To:")); dr.addWidget(self.d_to)
        g.addLayout(dr,3,1,1,3)
        g.addWidget(QLabel("Time:"),4,0)
        tr=QHBoxLayout(); tr.setSpacing(3)
        self.t_from=QTimeEdit(QTime(0,0,0)); self.t_from.setDisplayFormat("HH:mm:ss")
        self.t_to=QTimeEdit(QTime(23,59,59)); self.t_to.setDisplayFormat("HH:mm:ss")
        tr.addWidget(QLabel("From:")); tr.addWidget(self.t_from)
        tr.addWidget(QLabel("To:")); tr.addWidget(self.t_to)
        g.addLayout(tr,4,1,1,3)
        self.info_lbl=QLabel("")
        self.info_lbl.setStyleSheet("color:#8b949e;font-size:9px;")
        g.addWidget(self.info_lbl,5,0,1,4)
        if data:
            for attr,key,dflt in [("d_from","dt_from",QDate.currentDate().addDays(-1)),
                                   ("d_to","dt_to",QDate.currentDate())]:
                try:
                    d=datetime.datetime.strptime(data.get(key,"")[:10],"%Y-%m-%d")
                    getattr(self,attr).setDate(QDate(d.year,d.month,d.day))
                except: getattr(self,attr).setDate(dflt)
        else:
            self.d_from.setDate(QDate.currentDate().addDays(-1))
            self.d_to.setDate(QDate.currentDate())
        self.set_locked(locked)
    def set_locked(self,locked):
        self.locked=locked
        for w in [self.e_name,self.e_cells,self.d_from,self.d_to,self.t_from,self.t_to]:
            w.setEnabled(not locked)
        self.setStyleSheet(
            "#PC{background:#0f2318;border:2px solid #238636;border-radius:8px;margin:2px;}"
            if locked else
            "#PC{background:#161b22;border:1px solid #30363d;border-radius:8px;margin:2px;}")
    def get_data(self):
        cells=[c.strip() for c in self.e_cells.toPlainText().replace(",","\n").splitlines() if c.strip()]
        df2=self.d_from.date().toPyDate(); dt2=self.d_to.date().toPyDate()
        tf2=self.t_from.time().toPyTime(); tt2=self.t_to.time().toPyTime()
        return {"letter":self.letter,"name":self.e_name.text().strip(),"cells":cells,
                "dt_from":datetime.datetime.combine(df2,tf2).strftime("%Y-%m-%d %H:%M:%S"),
                "dt_to":  datetime.datetime.combine(dt2,tt2).strftime("%Y-%m-%d %H:%M:%S")}
    def validate(self):
        if not self.e_name.text().strip(): return f"Point {self.letter}: Name dein"
        if not [x for x in self.e_cells.toPlainText().replace(",","\n").splitlines() if x.strip()]:
            return f"Point {self.letter}: Cell IDs dein"
        return ""
    def set_info(self,msg,col="#8b949e"):
        self.info_lbl.setText(msg); self.info_lbl.setStyleSheet(f"color:{col};font-size:9px;")

# ─────────────────────────────────────────────────────────────────────────────
#  ROW BROWSER WORKER — loads all cases pkl and merges
# ─────────────────────────────────────────────────────────────────────────────
class RowBrowserWorker(QThread):
    progress=pyqtSignal(int,str); done=pyqtSignal(object); failed=pyqtSignal(str)
    def __init__(self,cases_info):
        # cases_info = list of (case_name, pkl_path)
        super().__init__(); self.cases_info=cases_info
    def run(self):
        try:
            all_dfs=[]; n=max(len(self.cases_info),1)
            for i,(cn,pkl) in enumerate(self.cases_info):
                self.progress.emit(int(i/n*90),f"Loading: {cn}")
                df=pkl_load(pkl)
                if df is not None and not df.empty:
                    df=df.copy()
                    df.insert(0,"Case_Name",cn)
                    all_dfs.append(df)
            if not all_dfs:
                self.done.emit(pd.DataFrame()); return
            merged=pd.concat(all_dfs,ignore_index=True)
            merged.insert(1,"#",range(1,len(merged)+1))
            self.progress.emit(100,"Done!"); self.done.emit(merged)
        except Exception as e:
            import traceback; self.failed.emit(f"{e}\n{traceback.format_exc()[:300]}")

# ─────────────────────────────────────────────────────────────────────────────
#  MAIN DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
class NexusDashboard(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("NEXUS ANALYZER PRO 8.0  |  Developed by Faisal Malik  |  Pakistan Forensic CDR Intelligence")
        self.setMinimumSize(1440,860); self.setStyleSheet(SS)
        self.base_path=BASE_PATH
        self.case_windows={}   # case_name -> CaseWindow
        self._rbw=None; self._row_df=pd.DataFrame()
        self._build_ui()
        self._refresh_cases()

    # ── UI Build ──────────────────────────────────────────────────────────────
    def _build_ui(self):
        cw=QWidget(); self.setCentralWidget(cw)
        root=QVBoxLayout(cw); root.setSpacing(0); root.setContentsMargins(0,0,0,0)

        # ── Header bar ─────────────────────────────────────────────────────────
        hb=QWidget(); hb.setFixedHeight(56)
        hb.setStyleSheet("background:#161b22;border-bottom:2px solid #21262d;")
        hl=QHBoxLayout(hb); hl.setContentsMargins(16,0,16,0); hl.setSpacing(12)
        hl.addWidget(_l(hb,"NEXUS ANALYZER PRO","#58a6ff",18,True))
        hl.addWidget(_l(hb,"v8.0","#8b949e",12))
        hl.addStretch()
        dev=QLabel("Developed by  Faisal Malik",hb)
        dev.setStyleSheet("color:#3fb950;font-size:12px;font-weight:bold;"
            "padding:5px 12px;border:1px solid #238636;border-radius:5px;background:#0d2a1a;")
        hl.addWidget(dev)
        hl.addWidget(_l(hb,"|","#30363d",14))
        hl.addWidget(_l(hb,"Pakistan Forensic CDR Intelligence","#8b949e",11))
        root.addWidget(hb)

        # ── Progress ───────────────────────────────────────────────────────────
        self._pb=QProgressBar(); self._pb.setFixedHeight(4); self._pb.setValue(0)
        self._pbl=_l(None,"","#8b949e",9)
        root.addWidget(self._pb); root.addWidget(self._pbl)

        # ── Main splitter: LEFT (create + cases) | RIGHT (browser) ─────────────
        spl=QSplitter(Qt.Horizontal); spl.setHandleWidth(1)

        # ── LEFT PANEL ─────────────────────────────────────────────────────────
        lw=QWidget(); lw.setFixedWidth(340)
        lw.setStyleSheet("background:#0d1117;border-right:1px solid #21262d;")
        ll=QVBoxLayout(lw); ll.setContentsMargins(10,10,10,10); ll.setSpacing(8)

        # Create new case
        cg=QGroupBox("Create New Case")
        cgl=QVBoxLayout(cg); cgl.setSpacing(6)
        cgl.addWidget(_l(cg,"Case Name:","#8b949e",10))
        self._cn_inp=QLineEdit(); self._cn_inp.setPlaceholderText("e.g. FIR-2024-001 Lahore")
        cgl.addWidget(self._cn_inp)
        cgl.addWidget(_l(cg,"Base Folder:","#8b949e",10))
        pr=QHBoxLayout(); pr.setSpacing(4)
        self._path_inp=QLineEdit(BASE_PATH); self._path_inp.setReadOnly(True)
        pr.addWidget(self._path_inp,1)
        pr.addWidget(_b("Browse","#21262d",self._browse_path,28))
        cgl.addLayout(pr)
        cgl.addWidget(_b("Create Case","#238636",self._create_case,32,True))
        ll.addWidget(cg)

        # Cases list
        csg=QGroupBox("Cases")
        csgl=QVBoxLayout(csg); csgl.setSpacing(5)
        sr=QHBoxLayout(); sr.setSpacing(4)
        self._case_search=QLineEdit(); self._case_search.setPlaceholderText("Search cases...")
        self._case_search.textChanged.connect(self._filter_cases)
        sr.addWidget(self._case_search,1)
        sr.addWidget(_b("Refresh","#21262d",self._refresh_cases,28))
        csgl.addLayout(sr)
        self._cases_list=QListWidget()
        self._cases_list.itemDoubleClicked.connect(self._open_case)
        csgl.addWidget(self._cases_list,1)
        br=QHBoxLayout(); br.setSpacing(4)
        br.addWidget(_b("Open Case","#1f6feb",self._open_case,30,True))
        br.addWidget(_b("Browse Folder","#21262d",self._browse_cases_dir,30))
        csgl.addLayout(br)
        ll.addWidget(csg,1)

        # Stats strip
        sg=QGroupBox("Dashboard Stats")
        sgl=QHBoxLayout(sg); sgl.setSpacing(8)
        self._ds_cases=self._stat_card("Cases","—","#58a6ff"); sgl.addWidget(self._ds_cases)
        self._ds_recs=self._stat_card("Records","—","#3fb950"); sgl.addWidget(self._ds_recs)
        self._ds_open=self._stat_card("Open","—","#e3b341"); sgl.addWidget(self._ds_open)
        ll.addWidget(sg)
        spl.addWidget(lw)

        # ── RIGHT PANEL: Row Browser ───────────────────────────────────────────
        rw=QWidget(); rl=QVBoxLayout(rw); rl.setContentsMargins(0,0,0,0); rl.setSpacing(0)

        # Browser toolbar
        bt=QWidget(); bt.setFixedHeight(46)
        bt.setStyleSheet("background:#0d1117;border-bottom:1px solid #21262d;")
        btl=QHBoxLayout(bt); btl.setContentsMargins(12,5,12,5); btl.setSpacing(8)
        btl.addWidget(_l(bt,"Row Browser — All Cases","#e3b341",13,True))
        btl.addWidget(_l(bt,"(Double-click row to open case)","#8b949e",10))
        btl.addStretch()
        # search
        self._rb_col=QComboBox()
        self._rb_col.addItems(["All","Case_Name","A-Party","B-Party","Cell ID","Network","IMEI","Location"])
        self._rb_col.setFixedWidth(110)
        self._rb_inp=QLineEdit(); self._rb_inp.setPlaceholderText("Search all CDR rows...")
        self._rb_inp.textChanged.connect(self._rb_search)
        rb_clr=QPushButton("X"); rb_clr.setFixedSize(24,26)
        rb_clr.setStyleSheet("background:#21262d;color:#e6edf3;border:none;border-radius:3px;")
        rb_clr.clicked.connect(lambda:self._rb_inp.clear())
        self._rb_cnt=QLabel(""); self._rb_cnt.setStyleSheet("color:#e3b341;font-size:10px;")
        btl.addWidget(self._rb_col); btl.addWidget(self._rb_inp,1)
        btl.addWidget(rb_clr); btl.addWidget(self._rb_cnt)
        self._rb_load=_b("Load All Cases","#6e40c9",self._load_row_browser,30,True)
        btl.addWidget(self._rb_load)
        rl.addWidget(bt)

        # Row browser tabs
        self._rb_tabs=QTabWidget()
        # Tab 1: Full row browser
        t1=QWidget(); t1l=QVBoxLayout(t1); t1l.setContentsMargins(2,2,2,2)
        self._rb_tv=make_tv()
        self._rb_tv.doubleClicked.connect(self._rb_open_case)
        t1l.addWidget(self._rb_tv)
        self._rb_tabs.addTab(t1,"All CDR Rows")
        # Tab 2: Cases summary
        t2=QWidget(); t2l=QVBoxLayout(t2); t2l.setContentsMargins(4,4,4,4)
        self._cases_tv=make_tv()
        t2l.addWidget(self._cases_tv)
        self._rb_tabs.addTab(t2,"Cases Summary")
        rl.addWidget(self._rb_tabs,1)

        # Log
        self._log_w=QTextEdit(); self._log_w.setFixedHeight(55); self._log_w.setReadOnly(True)
        self._log_w.setStyleSheet("background:#0d1117;border:none;border-top:1px solid #21262d;font-size:9px;")
        rl.addWidget(self._log_w)
        rl.addWidget(_footer_widget())

        spl.addWidget(rw); spl.setSizes([340,1100]); root.addWidget(spl,1)

    def _stat_card(self,label,val,col):
        f=QFrame(); f.setStyleSheet(f"QFrame{{background:#161b22;border:1px solid {col}33;"
            f"border-left:3px solid {col};border-radius:5px;padding:4px 8px;}}")
        fl=QVBoxLayout(f); fl.setSpacing(1); fl.setContentsMargins(0,0,0,0)
        v=QLabel(val); v.setStyleSheet(f"color:{col};font-size:16px;font-weight:bold;")
        lb=QLabel(label); lb.setStyleSheet("color:#8b949e;font-size:9px;")
        fl.addWidget(v); fl.addWidget(lb); f._v=v; return f

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _log(self,msg,col="#e6edf3"):
        ts=datetime.datetime.now().strftime("%H:%M:%S")
        self._log_w.append(f"<span style='color:#484f58'>[{ts}]</span> <span style='color:{col}'>{msg}</span>")
        self._log_w.verticalScrollBar().setValue(self._log_w.verticalScrollBar().maximum())

    def _prog(self,pct,msg=""):
        self._pb.setValue(max(0,min(pct,100)))
        if msg: self._pbl.setText(msg)

    def _cases_dir(self):
        p=self._path_inp.text().strip() or BASE_PATH
        return p

    def _case_path(self,cn):
        return os.path.join(self._cases_dir(),cn)

    def _meta_path(self,cn):
        return os.path.join(self._case_path(cn),"case_meta.json")

    def _pkl_for(self,cn):
        return os.path.join(self._case_path(cn),"Reports",f"{cn}.pkl")

    # ── Cases list ────────────────────────────────────────────────────────────
    def _refresh_cases(self):
        base=self._cases_dir()
        self._all_cases=[]
        if os.path.isdir(base):
            for name in sorted(os.listdir(base)):
                mp=os.path.join(base,name,"case_meta.json")
                if os.path.isfile(mp):
                    try:
                        with open(mp) as f: meta=json.load(f)
                    except: meta={"case_name":name}
                    self._all_cases.append((name,meta))
        self._display_cases(self._all_cases)
        self._refresh_cases_summary()

    def _display_cases(self,cases):
        self._cases_list.clear()
        for cn,meta in cases:
            recs=meta.get("records",0); pts=meta.get("points",[])
            ana=meta.get("last_analyzed","—")
            lock="LOCKED" if meta.get("locked",False) else "Open"
            txt=f"{cn}\n  {recs:,} records  |  {len(pts)} points  |  {lock}  |  {ana}"
            item=QListWidgetItem(txt)
            item.setData(Qt.UserRole,cn)
            if cn in self.case_windows:
                item.setForeground(QColor("#3fb950"))
            self._cases_list.addItem(item)
        self._ds_cases._v.setText(str(len(cases)))
        self._ds_open._v.setText(str(len(self.case_windows)))

    def _filter_cases(self,text):
        if not text:
            self._display_cases(self._all_cases); return
        filtered=[(cn,m) for cn,m in self._all_cases if text.lower() in cn.lower()]
        self._display_cases(filtered)

    def _refresh_cases_summary(self):
        rows=[]
        for cn,meta in self._all_cases:
            rows.append({
                "Case_Name":cn,
                "Records":meta.get("records",0),
                "Files":meta.get("file_count",0),
                "Points":len(meta.get("points",[])),
                "Status":"LOCKED" if meta.get("locked",False) else "Open",
                "Networks":meta.get("networks","—"),
                "Last_Analyzed":meta.get("last_analyzed","—"),
                "Has_Data":os.path.exists(self._pkl_for(cn)),
            })
        df=pd.DataFrame(rows)
        if not df.empty:
            tv_load(self._cases_tv,df)
            total_recs=df["Records"].sum()
            self._ds_recs._v.setText(f"{total_recs:,}")

    # ── Create / Browse ───────────────────────────────────────────────────────
    def _browse_path(self):
        p=QFileDialog.getExistingDirectory(self,"Select Base Folder",self._cases_dir())
        if p: self._path_inp.setText(p)

    def _browse_cases_dir(self):
        import subprocess
        p=self._cases_dir()
        if os.path.isdir(p):
            try: subprocess.Popen(["explorer",p])
            except: QMessageBox.information(self,"Path",p)

    def _create_case(self):
        cn=self._cn_inp.text().strip()
        if not cn: QMessageBox.warning(self,"Error","Case name dein."); return
        cn=re.sub(r'[\\/:*?"<>|]','_',cn)
        cp=self._case_path(cn)
        if os.path.exists(cp):
            QMessageBox.warning(self,"Exists",f"Case '{cn}' already exists."); return
        for sub in ["Raw_CDR","Reports"]:
            os.makedirs(os.path.join(cp,sub),exist_ok=True)
        meta={"case_name":cn,"created":datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
              "records":0,"file_count":0,"points":[],"locked":False}
        with open(os.path.join(cp,"case_meta.json"),"w") as f: json.dump(meta,f,indent=2)
        self._cn_inp.clear(); self._log(f"Case created: {cn}","#3fb950")
        self._refresh_cases()
        self._open_case_by_name(cn)

    def _open_case(self,item=None):
        if item is None:
            sel=self._cases_list.currentItem()
            if not sel: return
            item=sel
        cn=item.data(Qt.UserRole) if hasattr(item,'data') else item
        self._open_case_by_name(cn)

    def _open_case_by_name(self,cn):
        if cn in self.case_windows:
            w=self.case_windows[cn]; w.raise_(); w.activateWindow(); return
        mp=self._meta_path(cn)
        if not os.path.isfile(mp):
            QMessageBox.warning(self,"Error","Case folder not found."); return
        with open(mp) as f: meta=json.load(f)
        w=CaseWindow(self._case_path(cn),meta,self)
        w.closed.connect(self._on_case_closed)
        self.case_windows[cn]=w; w.show()
        self._refresh_cases()

    def _on_case_closed(self,cn):
        self.case_windows.pop(cn,None)
        self._refresh_cases()

    # ── Row Browser ───────────────────────────────────────────────────────────
    def _load_row_browser(self):
        cases_info=[]
        for cn,meta in self._all_cases:
            pkl=self._pkl_for(cn)
            if os.path.exists(pkl): cases_info.append((cn,pkl))
        if not cases_info:
            QMessageBox.information(self,"No Data","Koi case mein data import nahi hua."); return
        self._rb_load.setEnabled(False); self._prog(5,"Loading all cases...")
        self._log(f"Row Browser: loading {len(cases_info)} cases...","#8b949e")
        self._rbw=RowBrowserWorker(cases_info)
        self._rbw.progress.connect(self._prog)
        self._rbw.done.connect(self._rb_done)
        self._rbw.failed.connect(lambda m:(self._prog(0,""),self._log(f"Error: {m[:80]}","#f85149"),
            setattr(self,'_rb_load',None) or self._rb_load.setEnabled(True)))
        self._rbw.start()

    def _rb_done(self,df):
        self._row_df=df; self._rb_load.setEnabled(True); self._prog(0,"")
        n=len(df); self._rb_cnt.setText(f"{n:,} rows")
        disp=["Case_Name"]+[c for c in DISPLAY_COLS if c in df.columns]
        tv_load(self._rb_tv,df[disp])
        self._log(f"Row Browser loaded: {n:,} total rows from all cases","#3fb950")
        self._rb_tabs.setCurrentIndex(0)

    def _rb_search(self,text):
        if self._row_df.empty: return
        df=self._row_df
        disp=["Case_Name"]+[c for c in DISPLAY_COLS if c in df.columns]
        if not text.strip():
            tv_load(self._rb_tv,df[disp]); self._rb_cnt.setText(f"{len(df):,} rows"); return
        col=self._rb_col.currentText()
        if col=="All":
            mask=pd.Series(False,index=df.index)
            for c in disp: mask|=df[c].astype(str).str.contains(text,case=False,na=False,regex=False)
        else:
            if col not in df.columns: return
            mask=df[col].astype(str).str.contains(text,case=False,na=False,regex=False)
        result=df.loc[mask,disp]; tv_load(self._rb_tv,result); self._rb_cnt.setText(f"{len(result):,}")

    def _rb_open_case(self,idx):
        if self._row_df.empty: return
        row=self._row_df.iloc[idx.row()]
        cn=str(row.get("Case_Name",""))
        if cn: self._open_case_by_name(cn)

# ─────────────────────────────────────────────────────────────────────────────
#  CASE WINDOW  — one per case, full analysis workspace
# ─────────────────────────────────────────────────────────────────────────────
class CaseWindow(QMainWindow):
    closed=pyqtSignal(str)

    def __init__(self,case_path,case_meta,parent=None):
        super().__init__(parent)
        self.case_path=case_path; self.case_meta=dict(case_meta)
        self.master_df=pd.DataFrame(); self.analysis_dfs={}
        self.pts_locked=case_meta.get("locked",False)
        self.point_cards=[]; self.file_count=case_meta.get("file_count",0)
        self._common=set(); self._multi_im=set()
        self._iw=self._aw=self._clw=None
        cn=case_meta.get("case_name","Case")
        self.setWindowTitle(f"NEXUS — {cn}  |  Developed by Faisal Malik")
        self.setMinimumSize(1440,860); self.setStyleSheet(SS)
        self._build_ui(); self._load_data()

    def closeEvent(self,e):
        self.closed.emit(self.case_meta.get("case_name","")); super().closeEvent(e)

    # ── UI Build ──────────────────────────────────────────────────────────────
    def _build_ui(self):
        cw=QWidget(); self.setCentralWidget(cw)
        root=QVBoxLayout(cw); root.setSpacing(0); root.setContentsMargins(0,0,0,0)

        # ── Title bar ──────────────────────────────────────────────────────────
        tb=QWidget(); tb.setFixedHeight(52)
        tb.setStyleSheet("background:#161b22;border-bottom:2px solid #21262d;")
        tl=QHBoxLayout(tb); tl.setContentsMargins(12,0,12,0); tl.setSpacing(8)
        cn=self.case_meta.get("case_name","—")
        tl.addWidget(_l(tb,f"CASE:",  "#8b949e",11))
        tl.addWidget(_l(tb,cn,        "#58a6ff",14,True))
        self._recs_lbl=_l(tb,"","#8b949e",11); tl.addWidget(self._recs_lbl)
        tl.addStretch()
        dev=QLabel("Developed by Faisal Malik",tb)
        dev.setStyleSheet("color:#3fb950;font-size:10px;font-weight:bold;padding:3px 8px;"
                          "border:1px solid #238636;border-radius:4px;background:#0d2a1a;")
        tl.addWidget(dev); root.addWidget(tb)

        # ── Progress ───────────────────────────────────────────────────────────
        self._pb=QProgressBar(); self._pb.setFixedHeight(4); self._pb.setValue(0)
        self._pbl=_l(None,"","#8b949e",9)
        root.addWidget(self._pb); root.addWidget(self._pbl)

        # ── Info strip ─────────────────────────────────────────────────────────
        strip=QWidget(); strip.setFixedHeight(46)
        strip.setStyleSheet("background:#0d1117;border-bottom:1px solid #21262d;")
        sl=QHBoxLayout(strip); sl.setContentsMargins(12,3,12,3); sl.setSpacing(10)
        self._ic={
            "recs": _chip_widget(strip,"Records","—","#238636"),
            "files":_chip_widget(strip,"Files","—","#e3b341"),
            "pts":  _chip_widget(strip,"Points","—","#f0883e"),
            "lock": _chip_widget(strip,"Status","—","#8957e5"),
            "nets": _chip_widget(strip,"Networks","—","#58a6ff"),
            "susp": _chip_widget(strip,"Suspects","—","#f85149"),
            "drange":_chip_widget(strip,"Date Range","—","#79c0ff"),
        }
        for c in self._ic.values(): sl.addWidget(c)
        sl.addStretch(); root.addWidget(strip)

        # ── Main splitter ──────────────────────────────────────────────────────
        spl=QSplitter(Qt.Horizontal); spl.setHandleWidth(1)

        # ── LEFT SIDEBAR ───────────────────────────────────────────────────────
        lw=QWidget(); lw.setFixedWidth(310)
        lw.setStyleSheet("background:#0d1117;border-right:1px solid #21262d;")
        ll=QVBoxLayout(lw); ll.setContentsMargins(8,8,8,6); ll.setSpacing(5)

        # Import group
        ig=QGroupBox("CDR Import")
        igl=QVBoxLayout(ig); igl.setSpacing(5)
        ir=QHBoxLayout(); ir.setSpacing(4)
        self._b_imp=_b("Import CDR Files","#6e40c9",self._do_import,30,True)
        self._b_mas=_b("Export Master","#238636",self._export_master,30)
        self._b_mas.setEnabled(False)
        ir.addWidget(self._b_imp); ir.addWidget(self._b_mas); igl.addLayout(ir)
        self._imp_info=QLabel("No data"); self._imp_info.setStyleSheet("color:#8b949e;font-size:9px;")
        igl.addWidget(self._imp_info); ll.addWidget(ig)

        # Points group
        pg=QGroupBox("Crime Points  (Cell ID + Date/Time)")
        pgl=QVBoxLayout(pg); pgl.setSpacing(4); pgl.setContentsMargins(6,8,6,6)
        ph=QHBoxLayout(); ph.setSpacing(4)
        self._b_pt_add =_b("+ Add","#238636",self._pt_add,26)
        self._b_pt_save=_b("Save","#1f6feb",self._pt_save,26)
        self._b_pt_lock=_b("Lock","#6e40c9",self._pt_lock,26)
        for b in [self._b_pt_add,self._b_pt_save,self._b_pt_lock]: ph.addWidget(b)
        pgl.addLayout(ph)
        self._lk_lbl=QLabel("Edit mode"); self._lk_lbl.setStyleSheet("color:#238636;font-size:9px;")
        pgl.addWidget(self._lk_lbl)
        ptsc=QScrollArea(); ptsc.setWidgetResizable(True)
        self._pts_box=QWidget(); self._pts_lay=QVBoxLayout(self._pts_box)
        self._pts_lay.setSpacing(4); self._pts_lay.addStretch()
        ptsc.setWidget(self._pts_box); pgl.addWidget(ptsc,1); ll.addWidget(pg,1)

        # Analysis group
        ag=QGroupBox("Analysis & Export")
        agl=QVBoxLayout(ag); agl.setSpacing(4)
        self._b_ana=_b("Run Analysis","#1f6feb",self._do_analyze,30,True)
        self._b_ana.setEnabled(False)
        agl.addWidget(self._b_ana)
        er=QHBoxLayout(); er.setSpacing(4)
        self._b_exp_all=_b("Export All 10","#238636",self._export_all,28)
        self._b_exp_dos=_b("Export Dossier","#d2a8ff",self._export_dossier,28)
        er.addWidget(self._b_exp_all); er.addWidget(self._b_exp_dos); agl.addLayout(er)
        tr=QHBoxLayout(); tr.setSpacing(4)
        self._b_dtana =_b("Date/Time Analysis","#e3b341",self._open_dt_analysis,28)
        self._b_tower =_b("Tower Browser","#d2a8ff",self._open_tower_browser,28)
        tr.addWidget(self._b_dtana); tr.addWidget(self._b_tower); agl.addLayout(tr)
        ll.addWidget(ag)

        # Saved reports
        rg=QGroupBox("Saved Reports")
        rgl=QVBoxLayout(rg); rgl.setContentsMargins(4,6,4,4)
        self._saved_list=QListWidget(); self._saved_list.setMaximumHeight(130)
        self._saved_list.itemDoubleClicked.connect(self._view_report)
        rgl.addWidget(self._saved_list)
        rb=QHBoxLayout(); rb.setSpacing(4)
        rb.addWidget(_b("View","#1f6feb",self._view_report,24))
        rb.addWidget(_b("Folder","#21262d",self._open_folder,24))
        rgl.addLayout(rb); ll.addWidget(rg)
        spl.addWidget(lw)

        # ── RIGHT: Tabs ────────────────────────────────────────────────────────
        rw=QWidget(); rl=QVBoxLayout(rw); rl.setContentsMargins(0,0,0,0); rl.setSpacing(0)

        # Quick sheet bar
        bar=QWidget(); bar.setFixedHeight(36)
        bar.setStyleSheet("background:#0d1117;border-bottom:1px solid #21262d;")
        bl=QHBoxLayout(bar); bl.setContentsMargins(6,3,6,3); bl.setSpacing(3)
        self._sheet_btns={}
        for i,key in enumerate(SHEET_KEYS):
            ico,lbl2,bg,col=SHEET_META[key]
            b=QPushButton(f"{ico} {lbl2}"); b.setFixedHeight(26)
            b.setStyleSheet(f"background:{bg};color:{col}44;border:none;"
                           f"border-radius:3px;padding:1px 7px;font-size:10px;")
            b.clicked.connect(lambda _,i=i:self._tabs.setCurrentIndex(i))
            bl.addWidget(b); self._sheet_btns[key]=b
        bl.addStretch()
        bl.addWidget(_b("Export Sheet","#1f6feb",self._export_current,26))
        rl.addWidget(bar)

        # Main tabs
        self._tabs=QTabWidget()
        tab_defs=[
            ("Master CDR","_tv1"),("Common","_tv2"),("A-Party","_tv3"),
            ("IMEI","_tv4"),("Location","_tv5"),("Network","_tv6"),
            ("Direction","_tv7"),("Timeline","_tv8"),("Suspects","_tv9"),
            ("Intel","_tv10"),("DT Analysis","_tv11"),
        ]
        for lbl2,attr in tab_defs:
            t=QWidget(); tl2=QVBoxLayout(t); tl2.setContentsMargins(2,2,2,2)
            if attr=="_tv1":
                sr=QHBoxLayout(); sr.setSpacing(3)
                self._srch_col=QComboBox()
                self._srch_col.addItems(["All","A-Party","B-Party","Cell ID","Network","IMEI","Location","Site"])
                self._srch_col.setFixedWidth(100)
                self._srch_inp=QLineEdit(); self._srch_inp.setPlaceholderText("Search CDR data...")
                self._srch_inp.textChanged.connect(self._do_search)
                bx=QPushButton("X"); bx.setFixedSize(24,26)
                bx.setStyleSheet("background:#21262d;color:#e6edf3;border:none;border-radius:3px;")
                bx.clicked.connect(lambda:self._srch_inp.clear())
                self._srch_cnt=QLabel(""); self._srch_cnt.setStyleSheet("color:#e3b341;font-size:10px;")
                sr.addWidget(self._srch_col); sr.addWidget(self._srch_inp,1)
                sr.addWidget(bx); sr.addWidget(self._srch_cnt)
                tl2.addLayout(sr)
            tv=make_tv(); tl2.addWidget(tv); setattr(self,attr,tv)
            self._tabs.addTab(t,lbl2)
        try: self._tabs.tabBar().setTabTextColor(8,QColor("#f85149"))
        except: pass
        rl.addWidget(self._tabs,1)

        # Log + footer
        self._log_w=QTextEdit(); self._log_w.setFixedHeight(55); self._log_w.setReadOnly(True)
        self._log_w.setStyleSheet("background:#0d1117;border:none;border-top:1px solid #21262d;font-size:9px;")
        rl.addWidget(self._log_w)
        rl.addWidget(_footer_widget())
        spl.addWidget(rw); spl.setSizes([310,1130]); root.addWidget(spl,1)

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _log(self,msg,col="#e6edf3"):
        ts=datetime.datetime.now().strftime("%H:%M:%S")
        self._log_w.append(f"<span style='color:#484f58'>[{ts}]</span> <span style='color:{col}'>{msg}</span>")
        self._log_w.verticalScrollBar().setValue(self._log_w.verticalScrollBar().maximum())

    def _prog(self,pct,msg=""):
        self._pb.setValue(max(0,min(pct,100)))
        if msg: self._pbl.setText(msg)

    def _meta_save(self):
        with open(os.path.join(self.case_path,"case_meta.json"),"w") as f:
            json.dump(self.case_meta,f,indent=2)

    def _pkl_path(self):
        return os.path.join(self.case_path,"Reports",
                            f"{self.case_meta.get('case_name','case')}.pkl")

    def _update_strip(self):
        df=self.master_df; n=len(df) if not df.empty else 0
        self._ic["recs"]._v.setText(f"{n:,}" if n else "—")
        self._ic["files"]._v.setText(str(self.file_count))
        self._ic["pts"]._v.setText(str(len(self.case_meta.get("points",[]))))
        self._ic["lock"]._v.setText("LOCKED" if self.pts_locked else "Open")
        if "2_Common" in self.analysis_dfs and "A-Party" in self.analysis_dfs["2_Common"].columns:
            self._ic["susp"]._v.setText(str(self.analysis_dfs["2_Common"]["A-Party"].nunique()))
        if not df.empty and "Network" in df.columns:
            nv=df["Network"].value_counts()
            known=[x for x in nv.index if not str(x).startswith("Unknown")][:3]
            self._ic["nets"]._v.setText(", ".join(known) if known else "—")
        if not df.empty and "Date" in df.columns:
            dates=df["Date"].dropna(); dates=dates[dates!=""]
            if len(dates):
                self._ic["drange"]._v.setText(f"{dates.min()[:5]}..{dates.max()[:5]}")

    # ── Data load ─────────────────────────────────────────────────────────────
    def _load_data(self):
        for pt in self.case_meta.get("points",[]): self._add_pt_card(pt)
        self._update_lock_ui()
        cn=self.case_meta.get("case_name","")
        pkl=self._pkl_path()
        xl=os.path.join(self.case_path,"Reports",f"{cn}_Master.xlsx")
        self._prog(10,""); self._log("Loading...","#8b949e")
        self._clw=CaseLoadWorker(pkl,xl)
        self._clw.progress.connect(self._prog)
        self._clw.done.connect(self._on_loaded)
        self._clw.failed.connect(lambda m:self._log(f"Load: {m[:80]}","#f85149"))
        self._clw.start()

    def _on_loaded(self,df):
        self.master_df=df; self._prog(0,"")
        n=len(df) if not df.empty else 0
        if n:
            self._log(f"Loaded: {n:,} records","#238636")
            disp=[c for c in DISPLAY_COLS if c in df.columns]
            tv_load(self._tv1,df[disp])
            self._b_mas.setEnabled(True)
            self._recs_lbl.setText(f"{n:,} records")
            self._imp_info.setText(f"{n:,} records  |  {self.file_count} files")
        else:
            self._log("No data — Import CDR files","#8b949e")
        self._update_strip()
        can=self.pts_locked and not df.empty
        self._b_ana.setEnabled(can)
        self._refresh_saved()

    # ── Points ────────────────────────────────────────────────────────────────
    def _add_pt_card(self,data=None):
        card=PointCard(chr(65+len(self.point_cards)),data,self.pts_locked)
        card.sig_remove.connect(self._pt_remove)
        card.sig_browse.connect(self._open_cell_browser)
        self.point_cards.append(card)
        self._pts_lay.insertWidget(self._pts_lay.count()-1,card)

    def _pt_add(self):
        if self.pts_locked: QMessageBox.information(self,"Locked","Pehle unlock karein."); return
        if len(self.point_cards)>=26: QMessageBox.warning(self,"Max","Max 26 points."); return
        self._add_pt_card()

    def _pt_remove(self,card):
        if self.pts_locked: return
        self.point_cards.remove(card); card.setParent(None)
        for i,c in enumerate(self.point_cards):
            c.letter=chr(65+i); c.lbl.setText(f"Point {chr(65+i)}")

    def _pt_save(self):
        pts=[]
        for c in self.point_cards:
            e=c.validate()
            if e: QMessageBox.warning(self,"Error",e); return
            pts.append(c.get_data())
        self.case_meta["points"]=pts; self._meta_save()
        self._ic["pts"]._v.setText(str(len(pts)))
        self._log(f"{len(pts)} points saved","#238636")

    def _pt_lock(self):
        if self.pts_locked:
            if QMessageBox.question(self,"Unlock","Points unlock karein?",
               QMessageBox.Yes|QMessageBox.No)!=QMessageBox.Yes: return
            self.pts_locked=False
        else:
            for c in self.point_cards:
                e=c.validate()
                if e: QMessageBox.warning(self,"Error",e); return
            self._pt_save(); self.pts_locked=True
        self.case_meta["locked"]=self.pts_locked; self._meta_save()
        for c in self.point_cards: c.set_locked(self.pts_locked)
        self._update_lock_ui()
        self._b_ana.setEnabled(self.pts_locked and not self.master_df.empty)

    def _update_lock_ui(self):
        if self.pts_locked:
            self._lk_lbl.setText("LOCKED — analysis ready")
            self._lk_lbl.setStyleSheet("color:#f0883e;font-size:9px;font-weight:bold;")
            self._b_pt_lock.setText("Unlock")
            self._b_pt_lock.setStyleSheet("background:#6e1a1a;color:#fca5a5;"
                "border-radius:5px;padding:3px 8px;border:none;font-size:11px;")
        else:
            self._lk_lbl.setText("Edit mode — add points then Lock")
            self._lk_lbl.setStyleSheet("color:#238636;font-size:9px;")
            self._b_pt_lock.setText("Lock")
            self._b_pt_lock.setStyleSheet("background:#6e40c9;color:white;"
                "border-radius:5px;padding:3px 8px;border:none;font-size:11px;")

    # ── Search ────────────────────────────────────────────────────────────────
    def _do_search(self,text):
        if self.master_df.empty: return
        df=self.master_df; disp=[c for c in DISPLAY_COLS if c in df.columns]
        if not text.strip():
            tv_load(self._tv1,df[disp]); self._srch_cnt.setText(""); return
        col=self._srch_col.currentText()
        if col=="All":
            mask=pd.Series(False,index=df.index)
            for c in disp: mask|=df[c].astype(str).str.contains(text,case=False,na=False,regex=False)
        else:
            if col not in df.columns: return
            mask=df[col].astype(str).str.contains(text,case=False,na=False,regex=False)
        result=df.loc[mask,disp]; tv_load(self._tv1,result)
        self._srch_cnt.setText(f"{len(result):,}")

    # ── Import ────────────────────────────────────────────────────────────────
    def _do_import(self):
        files,_=QFileDialog.getOpenFileNames(self,"Select CDR Files",
            os.path.join(self.case_path,"Raw_CDR"),
            "CDR Files (*.csv *.xlsx *.xls *.txt);;All (*.*)")
        if not files: return
        self._b_imp.setEnabled(False); self._b_ana.setEnabled(False)
        self._prog(0,""); self._log(f"Importing {len(files)} file(s)...","#8b949e")
        raw=os.path.join(self.case_path,"Raw_CDR")
        self._iw=ImportWorker(files,self._pkl_path(),raw)
        self._iw.progress.connect(self._prog)
        self._iw.done.connect(self._import_done)
        self._iw.failed.connect(self._import_fail)
        self._iw.start()

    def _import_done(self,df,logs,fc):
        self.master_df=df; self.file_count+=fc
        self.case_meta["file_count"]=self.file_count
        self.case_meta["records"]=len(df)
        # Update networks in meta for dashboard
        if "Network" in df.columns:
            nv=df["Network"].value_counts()
            known=[x for x in nv.index if not str(x).startswith("Unknown")][:4]
            self.case_meta["networks"]=", ".join(known)
        self._meta_save()
        for l in logs: self._log(l,"#8b949e")
        n=len(df); self._log(f"Import complete: {n:,} records","#3fb950")
        self._b_imp.setEnabled(True); self._b_mas.setEnabled(True)
        self._imp_info.setText(f"{n:,} records  |  {self.file_count} files")
        self._recs_lbl.setText(f"{n:,} records")
        disp=[c for c in DISPLAY_COLS if c in df.columns]
        tv_load(self._tv1,df[disp]); self._update_strip(); self._prog(0,"")
        self._b_ana.setEnabled(self.pts_locked)
        if self.pts_locked:
            self._b_ana.setStyleSheet("background:#1f6feb;color:white;border-radius:5px;"
                "padding:3px 10px;border:none;font-weight:bold;font-size:12px;")

    def _import_fail(self,msg):
        self._b_imp.setEnabled(True); self._prog(0,"")
        QMessageBox.critical(self,"Import Error",msg[:400])
        self._log(f"Import error: {msg[:80]}","#f85149")

    def _export_master(self):
        if self.master_df.empty: return
        cn=self.case_meta["case_name"]
        fp=os.path.join(self.case_path,"Reports",f"{cn}_Master.xlsx")
        self._b_mas.setEnabled(False); self._prog(5,"Saving Master...")
        sw=SaveWorker([(self.master_df,fp)])
        sw.progress.connect(self._prog)
        sw.done.connect(lambda _:(self._prog(0,""),self._b_mas.setEnabled(True),
            self._log("Master.xlsx saved","#238636"),self._refresh_saved()))
        sw.failed.connect(lambda e:(self._prog(0,""),self._b_mas.setEnabled(True)))
        self._sw_m=sw; sw.start()

    # ── Analysis ──────────────────────────────────────────────────────────────
    def _do_analyze(self):
        if self.master_df.empty: QMessageBox.warning(self,"No Data","CDR import karein."); return
        if not self.pts_locked: QMessageBox.warning(self,"Points","Points lock karein."); return
        self._b_ana.setEnabled(False); self._b_imp.setEnabled(False)
        self._prog(0,""); self._log("Analysis shuru...","#8b949e")
        for key,b in self._sheet_btns.items():
            ico,lbl2,bg,col=SHEET_META[key]
            b.setStyleSheet(f"background:{bg};color:{col}33;border:none;"
                           f"border-radius:3px;padding:1px 7px;font-size:10px;")
        meta_wp=dict(self.case_meta); meta_wp["case_path"]=self.case_path
        self._aw=AnalysisWorker(self.master_df,meta_wp)
        self._aw.progress.connect(self._prog)
        self._aw.sheet.connect(self._on_sheet)
        self._aw.done.connect(self._analysis_done)
        self._aw.failed.connect(self._analysis_fail)
        self._aw.start()

    def _on_sheet(self,key,df):
        self.analysis_dfs[key]=df
        tv_map={"1_Master":self._tv1,"2_Common":self._tv2,"3_AParty":self._tv3,
                "4_IMEI":self._tv4,"5_Location":self._tv5,"6_Network":self._tv6,
                "7_Direction":self._tv7,"8_Timeline":self._tv8,"9_Suspects":self._tv9,
                "10_Intelligence":self._tv10}
        if key in tv_map: tv_load(tv_map[key],df)
        if key in self._sheet_btns:
            ico,lbl2,bg,col=SHEET_META[key]; b=self._sheet_btns[key]
            b.setStyleSheet(f"background:#0f2318;color:{col};border:1px solid {col}55;"
                           f"border-radius:3px;padding:1px 7px;font-size:10px;")
            self._log(f"{ico} {lbl2}: {len(df):,} rows","#3fb950")

    def _analysis_done(self,suspect_count,common,multi_im):
        self._common=common; self._multi_im=multi_im
        self._log(f"Analysis complete — {suspect_count} suspects","#f85149")
        self.case_meta["last_analyzed"]=datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        self._meta_save(); self._update_strip()
        self._prog(100,"Auto-saving..."); self._auto_save_bg()

    def _auto_save_bg(self):
        if not self.analysis_dfs: self._finalize(); return
        cn=self.case_meta.get("case_name","NEXUS")
        rdir=os.path.join(self.case_path,"Reports")
        ts=datetime.datetime.now().strftime("%Y%m%d_%H%M")
        SMAP={"1_Master":"1_Master","2_Common":"2_Common","3_AParty":"3_AParty",
              "4_IMEI":"4_IMEI","5_Location":"5_Location","6_Network":"6_Network",
              "7_Direction":"7_Direction","8_Timeline":"8_Timeline",
              "9_Suspects":"9_Suspects","10_Intelligence":"10_Intel"}
        tasks=[(self.analysis_dfs[k],os.path.join(rdir,f"{cn}_{sn}_{ts}.xlsx"))
               for k,sn in SMAP.items() if k in self.analysis_dfs]
        if not tasks: self._finalize(); return
        self._sw_a=SaveWorker(tasks)
        self._sw_a.progress.connect(self._prog)
        self._sw_a.done.connect(lambda _:(self._log(f"{len(tasks)} sheets saved","#238636"),
            self._finalize()))
        self._sw_a.failed.connect(lambda m:(self._log(f"Save warning: {m[:40]}","#f0883e"),
            self._finalize()))
        self._sw_a.start()

    def _finalize(self):
        self._b_ana.setEnabled(True); self._b_imp.setEnabled(True)
        has=not self.master_df.empty
        self._b_mas.setEnabled(has)
        self._prog(0,""); self._refresh_saved()

    def _analysis_fail(self,msg):
        self._b_ana.setEnabled(True); self._b_imp.setEnabled(True); self._prog(0,"")
        QMessageBox.critical(self,"Analysis Error",msg[:400])
        self._log(f"Analysis error: {msg[:80]}","#f85149")

    # ── Export ────────────────────────────────────────────────────────────────
    def _export_current(self):
        idx=self._tabs.currentIndex()
        keys=SHEET_KEYS; key=keys[idx] if idx<len(keys) else None
        if not key or key not in self.analysis_dfs:
            QMessageBox.warning(self,"No Data","Analysis run karein."); return
        cn=self.case_meta.get("case_name","NEXUS")
        fp,_=QFileDialog.getSaveFileName(self,"Save Sheet",
            os.path.join(self.case_path,"Reports",f"{cn}_{key}.xlsx"),
            "Excel Files (*.xlsx)")
        if not fp: return
        sw=SaveWorker([(self.analysis_dfs[key],fp)])
        sw.progress.connect(self._prog)
        sw.done.connect(lambda _:(self._prog(0,""),self._log(f"Sheet saved","#238636"),
            self._refresh_saved()))
        sw.failed.connect(lambda e:(self._prog(0,""),self._log(f"Export error","#f85149")))
        self._sw_c=sw; sw.start()

    def _export_all(self):
        if not self.analysis_dfs: QMessageBox.warning(self,"No Data","Analysis run karein."); return
        cn=self.case_meta.get("case_name","NEXUS")
        fp,_=QFileDialog.getSaveFileName(self,"Save 10-Sheet Report",
            os.path.join(self.case_path,"Reports",f"{cn}_Intelligence_Report.xlsx"),
            "Excel Files (*.xlsx)")
        if not fp: return
        SMAP={"1_Master":"1_Master","2_Common":"2_Common","3_AParty":"3_AParty",
              "4_IMEI":"4_IMEI","5_Location":"5_Location","6_Network":"6_Network",
              "7_Direction":"7_Direction","8_Timeline":"8_Timeline",
              "9_Suspects":"9_Suspects","10_Intelligence":"10_Intelligence"}
        dfs_map={sn:self.analysis_dfs[k] for k,sn in SMAP.items() if k in self.analysis_dfs}
        self._prog(5,f"Exporting {len(dfs_map)} sheets...")
        self._ew=MultiSheetExportWorker(dfs_map,fp)
        self._ew.progress.connect(self._prog)
        self._ew.done.connect(lambda _:(self._prog(0,""),
            self._log(f"10-Sheet report saved","#238636"),self._refresh_saved()))
        self._ew.failed.connect(lambda e:(self._prog(0,""),
            QMessageBox.critical(self,"Export Error",e[:300])))
        self._ew.start()

    def _export_dossier(self):
        if self.master_df.empty or not self._common:
            QMessageBox.warning(self,"No Data","Analysis run karein."); return
        cn=self.case_meta.get("case_name","NEXUS")
        fp,_=QFileDialog.getSaveFileName(self,"Save Intelligence Dossier",
            os.path.join(self.case_path,"Reports",f"{cn}_Intelligence_Dossier.xlsx"),
            "Excel Files (*.xlsx)")
        if not fp: return
        self._prog(5,"Writing Dossier...")
        self._dw=DossierWorker(self.master_df,self.case_meta,self._common,self._multi_im,fp)
        self._dw.progress.connect(self._prog)
        self._dw.done.connect(lambda _:(self._prog(0,""),
            self._log(f"Dossier saved","#d2a8ff"),self._refresh_saved()))
        self._dw.failed.connect(lambda e:(self._prog(0,""),
            QMessageBox.critical(self,"Dossier Error",e[:300])))
        self._dw.start()

    # ── Saved Reports ─────────────────────────────────────────────────────────
    def _refresh_saved(self):
        rdir=os.path.join(self.case_path,"Reports")
        self._saved_list.clear()
        if not os.path.isdir(rdir): return
        files=sorted([f for f in os.listdir(rdir) if f.endswith(".xlsx")],
                     key=lambda x:os.path.getmtime(os.path.join(rdir,x)),reverse=True)
        for fn in files:
            item=QListWidgetItem(fn)
            item.setData(Qt.UserRole,os.path.join(rdir,fn))
            self._saved_list.addItem(item)

    def _view_report(self):
        sel=self._saved_list.currentItem()
        if not sel: return
        fp=sel.data(Qt.UserRole)
        if not os.path.exists(fp): return
        try:
            df=pd.read_excel(fp,dtype=str,nrows=5000)
            if not df.empty:
                tv_load(self._tv1,df)
                self._tabs.setCurrentIndex(0)
                self._log(f"Viewing: {os.path.basename(fp)}","#8b949e")
        except Exception as e: self._log(f"View error: {e}","#f85149")

    def _open_folder(self):
        rdir=os.path.join(self.case_path,"Reports")
        os.makedirs(rdir,exist_ok=True)
        try:
            import subprocess; subprocess.Popen(["explorer",rdir])
        except: QMessageBox.information(self,"Folder",rdir)

    # ── Cell Browser for Point Cards ─────────────────────────────────────────
    def _open_cell_browser(self,card):
        if self.master_df.empty:
            QMessageBox.warning(self,"No Data","CDR import karein."); return
        df=self.master_df
        dlg=QDialog(self); dlg.setWindowTitle(f"Cell ID Browser — Point {card.letter}")
        dlg.setMinimumSize(900,520); dlg.setStyleSheet(SS)
        root=QVBoxLayout(dlg); root.setContentsMargins(10,10,10,10); root.setSpacing(6)
        root.addWidget(_l(dlg,"Cell ID Browser — Converted Values (Master Sheet)","#58a6ff",13,True))
        info=QLabel("Yeh values already converted hain (Telenor split, hex->dec). Yahi Point mein daalein.")
        info.setStyleSheet("background:#0d2a1a;color:#3fb950;border:1px solid #238636;"
                          "border-radius:4px;padding:5px;font-size:10px;")
        root.addWidget(info)
        fr=QHBoxLayout(); fr.setSpacing(6)
        fr.addWidget(QLabel("Network:"))
        net_cb=QComboBox()
        nets=["ALL"]+sorted(df["Network"].dropna().astype(str).unique().tolist())
        net_cb.addItems(nets); fr.addWidget(net_cb)
        fr.addWidget(QLabel("Search:"))
        srch=QLineEdit(); srch.setPlaceholderText("Cell ID / LAC / Site / Location...")
        fr.addWidget(srch,1); root.addLayout(fr)
        tv2=make_tv()
        tv2.setSelectionMode(QAbstractItemView.ExtendedSelection)
        root.addWidget(tv2,1)
        res_lbl=QLabel(""); res_lbl.setStyleSheet("color:#8b949e;font-size:10px;")
        root.addWidget(res_lbl)
        bb=QHBoxLayout()
        add_btn=QPushButton("Add to Point"); add_btn.setStyleSheet(
            "background:#238636;color:white;border:none;border-radius:5px;padding:5px 14px;font-weight:bold;")
        rep_btn=QPushButton("Replace Point"); rep_btn.setStyleSheet(
            "background:#e3b341;color:#0d1117;border:none;border-radius:5px;padding:5px 14px;font-weight:bold;")
        bb.addWidget(add_btn); bb.addWidget(rep_btn); bb.addStretch()
        bb.addWidget(_b("Close","#21262d",dlg.reject,28))
        root.addLayout(bb)
        # Build agg
        gc=[c for c in ["Cell ID","LAC","Network","Site","Location"] if c in df.columns]
        agg=(df.groupby(gc,dropna=False,sort=False).size().reset_index(name="Hits"))
        agg["Unique_Nums"]=df.groupby(gc,dropna=False,sort=False)["A-Party"].nunique().values
        agg=agg.sort_values("Hits",ascending=False).reset_index(drop=True)
        def populate(net="ALL",srch_txt=""):
            d=agg.copy()
            if net!="ALL": d=d[d["Network"].astype(str)==net]
            if srch_txt:
                mask=pd.Series(False,index=d.index)
                for c in gc: mask|=d[c].astype(str).str.contains(srch_txt,case=False,na=False,regex=False)
                d=d[mask]
            tv_load(tv2,d); res_lbl.setText(f"{len(d):,} towers")
        populate()
        net_cb.currentTextChanged.connect(lambda t:populate(t,srch.text()))
        srch.textChanged.connect(lambda t:populate(net_cb.currentText(),t))
        def get_sel():
            rows=set(idx.row() for idx in tv2.selectionModel().selectedRows())
            if not rows: return []
            mod=tv2.model()._df
            return [str(mod.iat[r,0]) for r in rows if 0<=r<len(mod)]
        def do_add():
            sel=get_sel()
            if not sel: return
            existing=card.e_cells.toPlainText().strip()
            new_cells="\n".join(([existing] if existing else [])+sel)
            card.e_cells.setPlainText(new_cells)
            card.set_info(f"{len(sel)} cells added","#3fb950"); dlg.accept()
        def do_replace():
            sel=get_sel()
            if not sel: return
            card.e_cells.setPlainText("\n".join(sel))
            card.set_info(f"{len(sel)} cells set","#e3b341"); dlg.accept()
        add_btn.clicked.connect(do_add); rep_btn.clicked.connect(do_replace)
        dlg.exec_()

    # ── Date/Time Analysis Window ─────────────────────────────────────────────
    def _open_dt_analysis(self):
        if self.master_df.empty: QMessageBox.warning(self,"No Data","CDR import karein."); return
        dlg=QDialog(self); dlg.setWindowTitle("Date / Time Range Analysis")
        dlg.setMinimumSize(1100,660); dlg.setStyleSheet(SS)
        root=QVBoxLayout(dlg); root.setSpacing(8); root.setContentsMargins(12,10,12,10)
        root.addWidget(_l(dlg,"Date / Time Range Cross-Match Analysis","#e3b341",14,True))
        root.addWidget(_l(dlg,"ABC/CBA Mutual Calls  •  IMEI Cross-Match  •  Active Days  •  Multi-Date","#8b949e",10))
        # Params
        pg2=QGroupBox("Parameters"); pgl=QHBoxLayout(pg2); pgl.setSpacing(8)
        pgl.addWidget(QLabel("Date From:")); dta_df=QDateEdit()
        dta_df.setCalendarPopup(True); dta_df.setDisplayFormat("dd/MM/yyyy")
        dta_df.setDate(QDate.currentDate().addDays(-7)); pgl.addWidget(dta_df)
        pgl.addWidget(QLabel("To:")); dta_dt=QDateEdit()
        dta_dt.setCalendarPopup(True); dta_dt.setDisplayFormat("dd/MM/yyyy")
        dta_dt.setDate(QDate.currentDate()); pgl.addWidget(dta_dt)
        pgl.addWidget(QLabel("Time From:"))
        dta_tf=QTimeEdit(QTime(0,0,0)); dta_tf.setDisplayFormat("HH:mm:ss"); pgl.addWidget(dta_tf)
        pgl.addWidget(QLabel("To:"))
        dta_tt=QTimeEdit(QTime(23,59,59)); dta_tt.setDisplayFormat("HH:mm:ss"); pgl.addWidget(dta_tt)
        pgl.addWidget(QLabel("Network:"))
        dta_net=QComboBox(); dta_net.setFixedWidth(90)
        dta_net.addItems(["All"]+sorted(self.master_df["Network"].dropna().astype(str).unique().tolist()))
        pgl.addWidget(dta_net)
        pgl.addWidget(QLabel("A-Party:"))
        dta_ap=QLineEdit(); dta_ap.setFixedWidth(110); dta_ap.setPlaceholderText("e.g. 0300")
        pgl.addWidget(dta_ap); pgl.addStretch()
        dta_pb=QProgressBar(); dta_pb.setFixedWidth(130); dta_pb.setFixedHeight(6); pgl.addWidget(dta_pb)
        dta_run=QPushButton("RUN"); dta_run.setFixedSize(58,30)
        dta_run.setStyleSheet("background:#e3b341;color:#0d1117;border:none;border-radius:5px;font-weight:bold;")
        pgl.addWidget(dta_run); root.addWidget(pg2)
        res_lbl=QLabel("Set parameters and click RUN")
        res_lbl.setStyleSheet("color:#8b949e;font-size:10px;"); root.addWidget(res_lbl)
        res_tv=make_tv(); root.addWidget(res_tv,1)
        bb=QHBoxLayout()
        exp_btn=QPushButton("Export Results")
        exp_btn.setStyleSheet("background:#238636;color:white;border:none;border-radius:5px;padding:5px 14px;")
        exp_btn.setEnabled(False); bb.addWidget(exp_btn); bb.addStretch()
        bb.addWidget(_b("Close","#21262d",dlg.reject,28)); root.addLayout(bb)
        holder=[None]
        def run_dta():
            f1=dta_df.date().toPyDate().strftime("%Y-%m-%d")
            f2=dta_dt.date().toPyDate().strftime("%Y-%m-%d")
            tf=dta_tf.time().toString("HH:mm:ss"); tt=dta_tt.time().toString("HH:mm:ss")
            net=dta_net.currentText(); ap=dta_ap.text().strip()
            flt={}
            if net!="All": flt["Network"]=net
            if ap: flt["A-Party"]=ap
            dta_run.setEnabled(False); dta_pb.setValue(10)
            res_lbl.setText("Running..."); res_lbl.setStyleSheet("color:#8b949e;font-size:10px;")
            w=DTAnalysisWorker(self.master_df,f1,f2,tf,tt,flt)
            w.progress.connect(dta_pb.setValue)
            def on_done(df):
                dta_run.setEnabled(True); holder[0]=df; tv_load(res_tv,df)
                mutual=(df["ABC_CBA"]=="MUTUAL").sum() if "ABC_CBA" in df.columns else 0
                multi=(df["Multi_IMEI"]=="YES").sum() if "Multi_IMEI" in df.columns else 0
                res_lbl.setText(f"{len(df):,} numbers  |  {mutual} ABC/CBA mutual  |  {multi} multi-IMEI  |  {f1} to {f2}")
                res_lbl.setStyleSheet("color:#3fb950;font-size:10px;font-weight:bold;")
                exp_btn.setEnabled(True); dta_pb.setValue(100)
            def on_fail(msg):
                dta_run.setEnabled(True); dta_pb.setValue(0)
                res_lbl.setText(f"Error: {msg[:80]}")
                res_lbl.setStyleSheet("color:#f85149;")
            w.done.connect(on_done); w.failed.connect(on_fail); self._dta_w=w; w.start()
        def export_dta():
            if holder[0] is None: return
            fp2,_=QFileDialog.getSaveFileName(dlg,"Save Analysis",
                os.path.join(self.case_path,"Reports","DateTime_Analysis.xlsx"),
                "Excel (*.xlsx)")
            if not fp2: return
            exp_btn.setEnabled(False)
            sw=SaveWorker([(holder[0],fp2)])
            sw.progress.connect(dta_pb.setValue)
            sw.done.connect(lambda _:(exp_btn.setEnabled(True),dta_pb.setValue(0),
                self._log(f"DT Analysis saved","#e3b341"),self._refresh_saved()))
            sw.failed.connect(lambda e:exp_btn.setEnabled(True))
            self._dta_sw=sw; sw.start()
        dta_run.clicked.connect(run_dta); exp_btn.clicked.connect(export_dta)
        dlg.exec_()

    # ── Tower Browser ─────────────────────────────────────────────────────────
    def _open_tower_browser(self):
        if self.master_df.empty: QMessageBox.warning(self,"No Data","CDR import karein."); return
        df=self.master_df
        dlg=QDialog(self); dlg.setWindowTitle("Tower / Cell ID Browser")
        dlg.setMinimumSize(1050,620); dlg.setStyleSheet(SS)
        root=QVBoxLayout(dlg); root.setContentsMargins(10,10,10,10); root.setSpacing(6)
        root.addWidget(_l(dlg,"Tower / Cell ID Browser","#d2a8ff",14,True))
        root.addWidget(_l(dlg,"Converted Cell IDs & LAC values — search, filter, copy to clipboard","#8b949e",10))
        fr=QHBoxLayout(); fr.setSpacing(8)
        fr.addWidget(QLabel("Network:"))
        net_cb=QComboBox(); net_cb.setFixedWidth(100)
        net_cb.addItems(["ALL"]+sorted(df["Network"].dropna().astype(str).unique().tolist()))
        fr.addWidget(net_cb)
        fr.addWidget(QLabel("Search:"))
        srch=QLineEdit(); srch.setPlaceholderText("Cell ID / LAC / Site / Location...")
        fr.addWidget(srch,1)
        cnt_lbl=QLabel(""); cnt_lbl.setStyleSheet("color:#e3b341;font-size:11px;font-weight:bold;")
        fr.addWidget(cnt_lbl); root.addLayout(fr)
        tv3=make_tv(); root.addWidget(tv3,1)
        # Build aggregated tower data
        gc=[c for c in ["Cell ID","LAC","Network","Site","Location"] if c in df.columns]
        agg=(df.groupby(gc,dropna=False,sort=False).agg(
            Hits=("A-Party","count"),
            Unique_Numbers=("A-Party","nunique"),
            Active_Days=("Date","nunique"),
            First_Date=("Date","min"),
            Last_Date=("Date","max")).reset_index()
             .sort_values("Hits",ascending=False).reset_index(drop=True))
        # Add suspect count
        if self._common:
            sc=df[df["A-Party"].isin(self._common)].groupby("Cell ID",sort=False)["A-Party"].nunique()
            agg["Suspect_Count"]=agg["Cell ID"].map(sc).fillna(0).astype(int)
        def populate(net="ALL",txt=""):
            d=agg.copy()
            if net!="ALL": d=d[d["Network"].astype(str)==net]
            if txt:
                mask=pd.Series(False,index=d.index)
                for c in gc: mask|=d[c].astype(str).str.contains(txt,case=False,na=False,regex=False)
                d=d[mask]
            tv_load(tv3,d); cnt_lbl.setText(f"{len(d):,} towers")
        populate()
        net_cb.currentTextChanged.connect(lambda t:populate(t,srch.text()))
        srch.textChanged.connect(lambda t:populate(net_cb.currentText(),t))
        bb=QHBoxLayout()
        def copy_cells():
            rows=set(idx.row() for idx in tv3.selectionModel().selectedRows())
            if not rows:
                QMessageBox.information(dlg,"Select","Rows select karein."); return
            mod=tv3.model()._df
            cells="\n".join(str(mod.iat[r,0]) for r in sorted(rows) if 0<=r<len(mod))
            QApplication.clipboard().setText(cells)
            QMessageBox.information(dlg,"Copied",f"{len(rows)} Cell IDs clipboard mein copy ho gaye.")
        def export_tw():
            fp2,_=QFileDialog.getSaveFileName(dlg,"Save Tower Data",
                os.path.join(self.case_path,"Reports","Tower_Browser.xlsx"),"Excel (*.xlsx)")
            if not fp2: return
            sw=SaveWorker([(agg,fp2)])
            sw.progress.connect(lambda p,m:None)
            sw.done.connect(lambda _:(self._log(f"Tower data exported","#d2a8ff"),
                self._refresh_saved()))
            self._tw_sw=sw; sw.start()
        bb.addWidget(_b("Copy Selected Cell IDs","#58a6ff",copy_cells,30,True))
        bb.addWidget(_b("Export All","#238636",export_tw,30))
        bb.addStretch()
        bb.addWidget(_b("Close","#21262d",dlg.reject,30))
        root.addLayout(bb); dlg.exec_()

# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────
if __name__=="__main__":
    app=QApplication(sys.argv); app.setStyle("Fusion")
    win=NexusDashboard(); win.show(); sys.exit(app.exec_())
